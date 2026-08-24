from fastapi import FastAPI, Form, Request, UploadFile, File, Depends, Query
from starlette.middleware.sessions import SessionMiddleware
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, case
from sqlalchemy.orm import sessionmaker, declarative_base
from jinja2 import Environment, FileSystemLoader
from sqlalchemy import text
import pandas as pd
import io
from datetime import datetime
from fastapi.responses import FileResponse
import pdfkit
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from fastapi.responses import HTMLResponse
from database import Base 

from jinja2 import Environment, FileSystemLoader
import os
from chatbot_agent import query_chatbot
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


# ---------------- PROCESS ORDER ----------------

app = FastAPI()

# ── Session middleware ──
app.add_middleware(SessionMiddleware, secret_key="inventory-secret-key-2026")

# ── Hardcoded credentials ──
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"

# ── Stores configuration ──
STORES = {
    "mahindra": "Mahindra Pro Spares",
    "divya": "Divya Automobiles"
}

def get_active_store(request: Request) -> str:
    store = request.session.get("active_store")
    if store not in STORES:
        return "mahindra"
    return store

@app.get("/select_store", response_class=HTMLResponse)
def select_store_page(request: Request):
    if not request.session.get("user"):
        return RedirectResponse("/login", status_code=302)
    active_store = request.session.get("active_store", "")
    return render("select_store.html", stores=STORES, active_store=active_store)

@app.get("/switch_store/{store_name}")
def switch_store(store_name: str, request: Request):
    if store_name in STORES:
        request.session["active_store"] = store_name
        return RedirectResponse("/", status_code=302)
    return RedirectResponse("/select_store", status_code=302)

# ── Auto-import inventory.xlsx on startup if products table is empty ──
@app.on_event("startup")
def auto_import_inventory():
    db = SessionLocal()
    try:
        from sqlalchemy import text as _text

        # ── 0. Run auto-migrations for invoices & quotations ──
        for table in ["invoices", "quotations"]:
            for col, col_type in [
                ("customer_address", "VARCHAR(255)"),
                ("customer_gstin", "VARCHAR(50)"),
                ("customer_mobile", "VARCHAR(50)"),
                ("customer_email", "VARCHAR(255)")
            ]:
                try:
                    db.execute(_text(f"ALTER TABLE {table} ADD COLUMN {col} {col_type}"))
                    db.commit()
                except Exception:
                    db.rollback()

        # Add purchase_rate to invoice_items
        try:
            db.execute(_text("ALTER TABLE invoice_items ADD COLUMN purchase_rate FLOAT DEFAULT 0.0"))
            db.commit()
        except Exception:
            db.rollback()

        # Add vendor details, location, and store column to products table & others
        for tbl in ["products", "invoices", "quotations", "purchases", "order_items", "vendors"]:
            try:
                db.execute(_text(f"ALTER TABLE {tbl} ADD COLUMN store VARCHAR(50) DEFAULT 'mahindra'"))
                db.commit()
            except Exception:
                db.rollback()
            try:
                db.execute(_text(f"UPDATE {tbl} SET store = 'mahindra' WHERE store IS NULL OR store = ''"))
                db.commit()
            except Exception:
                db.rollback()

        for col, col_type in [
            ("vendor_name", "VARCHAR(255)"),
            ("vendor_address", "VARCHAR(255)"),
            ("vendor_mobile", "VARCHAR(50)"),
            ("vendor_gstin", "VARCHAR(50)"),
            ("vendor_email", "VARCHAR(255)"),
            ("location", "VARCHAR(100)")
        ]:
            try:
                db.execute(_text(f"ALTER TABLE products ADD COLUMN {col} {col_type}"))
                db.commit()
            except Exception:
                db.rollback()

        # ── 1. Import inventory.xlsx → products table ──
        count = db.execute(_text("SELECT COUNT(*) FROM products")).scalar()
        if count == 0:
            xlsx_path = os.path.join(os.path.dirname(__file__), "inventory.xlsx")
            if os.path.exists(xlsx_path):
                import pandas as pd
                df = pd.read_excel(xlsx_path, engine="openpyxl")
                df.columns = [col.strip() for col in df.columns]
                for _, row in df.iterrows():
                    qty      = int(row.get("Qty", 0) or 0)
                    rate     = float(row.get("Rate", 0) or 0)
                    discount = float(row.get("Discount", 0) or 0)
                    amount   = ((qty * rate) - (qty * rate * discount / 100)) if qty > 0 else 0.0
                    loc = str(row.get("Location", row.get("location", "")) or "").strip()
                    db.execute(_text("""
                        INSERT INTO products (part_no, description, hsn, gst, quantity, rate, discount, amount, location, store)
                        VALUES (:pn, :desc, :hsn, :gst, :qty, :rate, :disc, :amt, :loc, 'mahindra')
                    """), {
                        "pn":   str(row.get("Part No", "")).strip(),
                        "desc": str(row.get("Description", "")).strip(),
                        "hsn":  str(row.get("HSN", "")).strip(),
                        "gst":  float(row.get("GST", 18) or 18),
                        "qty":  qty, "rate": rate, "disc": discount,
                        "amt":  round(amount, 2),
                        "loc":  loc
                    })
                db.commit()
                print(f"✅ Auto-imported inventory.xlsx ({len(df)} products)")

        # ── 2. Sync orders.csv → order_items table (always refresh on startup) ──
        csv_path = os.path.join(os.path.dirname(__file__), "orders.csv")
        if os.path.exists(csv_path):
            oi_count = db.execute(_text("SELECT COUNT(*) FROM order_items")).scalar()
            import pandas as pd
            df_csv = pd.read_csv(csv_path, dtype=str, on_bad_lines="skip")
            df_csv.columns = [col.strip() for col in df_csv.columns]
            csv_count = len(df_csv)
            # Only reimport if DB is out of sync with CSV (allows Render cold starts)
            if oi_count < csv_count * 0.9:  # reimport if DB has <90% of CSV rows
                print(f"⏳ Syncing orders.csv ({csv_count} rows) into order_items ({oi_count} in DB)...")
                db.execute(_text("DELETE FROM order_items"))
                db.commit()
                batch = []
                for _, row in df_csv.iterrows():
                    part_no = str(row.get("Part No", "") or "").strip()
                    if not part_no:
                        continue
                    batch.append({
                        "pn":   part_no,
                        "desc": str(row.get("Part Desc", "") or "").strip(),
                        "hsn":  str(row.get("HSN", "") or "").strip(),
                        "mrp":  float(row.get("MRP", 0) or 0)
                    })
                    if len(batch) >= 500:
                        params = {}
                        values_clauses = []
                        for i, item in enumerate(batch):
                            values_clauses.append(f"(:pn_{i}, :desc_{i}, :hsn_{i}, :mrp_{i}, 'mahindra')")
                            params[f"pn_{i}"] = item["pn"]
                            params[f"desc_{i}"] = item["desc"]
                            params[f"hsn_{i}"] = item["hsn"]
                            params[f"mrp_{i}"] = item["mrp"]
                        sql = f"INSERT INTO order_items (part_no, description, hsn, mrp, store) VALUES {', '.join(values_clauses)}"
                        db.execute(_text(sql), params)
                        db.commit()
                        batch = []
                if batch:
                    params = {}
                    values_clauses = []
                    for i, item in enumerate(batch):
                        values_clauses.append(f"(:pn_{i}, :desc_{i}, :hsn_{i}, :mrp_{i}, 'mahindra')")
                        params[f"pn_{i}"] = item["pn"]
                        params[f"desc_{i}"] = item["desc"]
                        params[f"hsn_{i}"] = item["hsn"]
                        params[f"mrp_{i}"] = item["mrp"]
                    sql = f"INSERT INTO order_items (part_no, description, hsn, mrp, store) VALUES {', '.join(values_clauses)}"
                    db.execute(_text(sql), params)
                    db.commit()
                print(f"✅ Synced orders.csv ({csv_count} products into order_items)")
            else:
                print(f"✅ order_items up to date ({oi_count} rows)")
        else:
            print("⚠ orders.csv not found — skipping order_items sync")

    except Exception as e:
        print(f"⚠ Auto-import skipped: {e}")
        import traceback; traceback.print_exc()
    finally:
        db.close()

# ── Auth helper ──
def get_current_user(request: Request):
    return request.session.get("user")
# Raw Jinja2 rendering — bypasses Starlette Jinja2Templates API version issues
_jinja_env = Environment(
    loader=FileSystemLoader("templates"),
    cache_size=0,
    auto_reload=True
)

def render(name: str, status_code: int = 200, **ctx) -> HTMLResponse:
    html = _jinja_env.get_template(name).render(**ctx)
    return HTMLResponse(content=html, status_code=status_code)
from database import Base, engine, SessionLocal

Base.metadata.create_all(bind=engine)

# ---------------- MODELS ----------------

class Product(Base):
    __tablename__ = "products"
    id = Column(Integer, primary_key=True)
    part_no = Column(String(100))
    description = Column(String(255))
    hsn = Column(String(50))
    gst = Column(Float)
    quantity = Column(Integer)
    rate = Column(Float)
    discount = Column(Float)
    amount = Column(Float)
    location = Column(String(100), nullable=True, default="")
    store = Column(String(50), default="mahindra", index=True)
    vendor_name = Column(String(255), nullable=True)
    vendor_address = Column(String(255), nullable=True)
    vendor_mobile = Column(String(50), nullable=True)
    vendor_gstin = Column(String(50), nullable=True)
    vendor_email = Column(String(255), nullable=True)

class Invoice(Base):
    __tablename__ = "invoices"

    id = Column(Integer, primary_key=True)
    invoice_no = Column(String(50), unique=True)
    customer_name = Column(String(255))
    customer_address = Column(String(255), nullable=True)
    customer_gstin = Column(String(50), nullable=True)
    customer_mobile = Column(String(50), nullable=True)
    customer_email = Column(String(255), nullable=True)
    date = Column(DateTime, default=datetime.utcnow)
    store = Column(String(50), default="mahindra", index=True)

    total_amount = Column(Float)
    gst_amount = Column(Float)
    grand_total = Column(Float)

class InvoiceItem(Base):
    __tablename__ = "invoice_items"

    id = Column(Integer, primary_key=True)

    invoice_id = Column(Integer)  # FK (can add relationship later)

    part_no = Column(String(100))
    description = Column(String(255))

    quantity = Column(Integer)
    rate = Column(Float)
    amount = Column(Float)

    hsn = Column(String(50))
    purchase_rate = Column(Float, default=0.0)

class OrderItems(Base):
    __tablename__ = "order_items"

    id = Column(Integer, primary_key=True)
    part_no = Column(String(100))
    description = Column(String(255))
    hsn = Column(String(50))
    mrp = Column(Float)
    store = Column(String(50), default="mahindra", index=True)

class Quotation(Base):
    __tablename__ = "quotations"
    id = Column(Integer, primary_key=True)
    quotation_no = Column(String(50))
    customer_name = Column(String(255), default="")
    customer_address = Column(String(255), default="")
    customer_gstin = Column(String(50), default="")
    customer_mobile = Column(String(50), default="")
    customer_email = Column(String(255), default="")
    date = Column(DateTime, default=datetime.utcnow)
    total_amount = Column(Float, default=0.0)
    grand_total = Column(Float, default=0.0)
    store = Column(String(50), default="mahindra", index=True)

class QuotationItem(Base):
    __tablename__ = "quotation_items"
    id = Column(Integer, primary_key=True)
    quotation_id = Column(Integer)
    part_no = Column(String(100))
    description = Column(String(255))
    rate = Column(Float)
    qty = Column(Integer)
    discount = Column(Float)
    amount = Column(Float)
    hsn = Column(String(50), default="")

class Vendor(Base):
    __tablename__ = "vendors"
    id = Column(Integer, primary_key=True)
    name = Column(String(255))
    address = Column(String(255), nullable=True)
    mobile_num = Column(String(50), nullable=True)
    gstin = Column(String(50), nullable=True)
    email_id = Column(String(255), nullable=True)
    store = Column(String(50), default="mahindra", index=True)

class Purchase(Base):
    __tablename__ = "purchases"
    id = Column(Integer, primary_key=True)
    vendor_name = Column(String(255))
    part_no = Column(String(100))
    description = Column(String(255))
    hsn = Column(String(50))
    quantity = Column(Integer)
    rate = Column(Float)
    discount = Column(Float)
    amount = Column(Float)
    date = Column(DateTime, default=datetime.utcnow)
    store = Column(String(50), default="mahindra", index=True)

Base.metadata.create_all(bind=engine)

def format_inr(amount):
    s = f"{amount:,.2f}"
    parts = s.split(".")
    integer = parts[0]
    decimal = parts[1]

    if len(integer) > 3:
        last3 = integer[-3:]
        rest = integer[:-3]
        rest = ",".join([rest[max(i-2,0):i] for i in range(len(rest), 0, -2)][::-1])
        integer = rest + "," + last3

    return f"₹ {integer}.{decimal}"

# ---------------- HOME ----------------

# @app.get("/", response_class=HTMLResponse)
# def home(request: Request):
#     db = SessionLocal()
#     products = db.query(Product).all()

#     products = sorted(products, key=lambda x: x.quantity > 0)

#     # 🔥 ADD THIS
#     total_value = sum([(p.amount or 0) for p in products])
#     formatted_total = format_inr(total_value)

#     db.close()

#     return templates.TemplateResponse(
#         "index.html",
#         context={
#             "request": request,
#             "products": products,
#             "total_value": round(total_value, 2)
#         }
#     )

# ── Login routes ──
@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if request.session.get("user"):
        if not get_active_store(request):
            return RedirectResponse("/select_store", status_code=302)
        return RedirectResponse("/", status_code=302)
    return render("login.html", error=None)

@app.post("/login", response_class=HTMLResponse)
def login(request: Request, username: str = Form(...), password: str = Form(...)):
    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        request.session["user"] = username
        return RedirectResponse("/select_store", status_code=303)
    return render("login.html", status_code=401, error="Invalid username or password. Please try again.")

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=302)

# ── Protected home route ──
@app.get("/", response_class=HTMLResponse)
def home(request: Request, page: int = 1, q: str = ""):
    if not request.session.get("user"):
        return RedirectResponse("/login", status_code=302)

    active_store = get_active_store(request)
    if not active_store:
        return RedirectResponse("/select_store", status_code=302)

    store_name_display = STORES.get(active_store, "Mahindra Pro Spares")

    db = SessionLocal()
    from sqlalchemy import func

    per_page = 50
    offset = (page - 1) * per_page
    q_clean = q.strip().upper()

    query = db.query(Product).filter(Product.store == active_store)
    if q_clean:
        query = query.filter(
            (func.upper(Product.part_no).like(f"%{q_clean}%")) |
            (func.upper(Product.description).like(f"%{q_clean}%")) |
            (func.upper(Product.location).like(f"%{q_clean}%"))
        )

    ordered_query = query.order_by(
        case(
            (Product.quantity == 0, 0),
            (Product.quantity < 5, 1),
            else_=2
        )
    )

    total_items = ordered_query.count()
    products = ordered_query.offset(offset).limit(per_page).all()

    # Total value of all stock in DB for active store
    total_value_query = db.query(Product).filter(Product.store == active_store, Product.quantity > 0).all()
    total_value = sum(
        (p.quantity * (p.rate or 0)) - ((p.quantity * (p.rate or 0)) * ((p.discount or 0) / 100))
        for p in total_value_query
    )

    # Global stats for active store dashboard
    stat_total = db.query(Product).filter(Product.store == active_store).count()
    stat_instock = db.query(Product).filter(Product.store == active_store, Product.quantity > 0).count()
    stat_outstock = db.query(Product).filter(Product.store == active_store, Product.quantity == 0).count()

    total_pages = (total_items + per_page - 1) // per_page
    if total_pages < 1:
        total_pages = 1

    db.close()

    return render(
        "index.html",
        products=products,
        total_value=round(total_value, 2),
        current_page=page,
        total_pages=total_pages,
        total_items=total_items,
        stat_total=stat_total,
        stat_instock=stat_instock,
        stat_outstock=stat_outstock,
        q=q,
        active_store=active_store,
        store_name_display=store_name_display,
        stores=STORES
    )
# ---------------- PRICE MASTER UPLOAD ----------------


from fastapi.responses import StreamingResponse
import pandas as pd
import io

@app.get("/export_excel")
def export_excel(
    request: Request,
    start_date: str,
    end_date: str,
    customer: str = Query(""),
    vendor: str = Query("")
):
    active_store = get_active_store(request)
    db = SessionLocal()

    sql_parts = [
        "FROM invoice_items ii",
        "JOIN invoices i ON ii.invoice_id = i.id",
        "LEFT JOIN products p ON (ii.part_no = p.part_no AND p.store = :store)",
        "WHERE i.store = :store AND CAST(i.date AS DATE) BETWEEN :start AND :end"
    ]
    
    params = {
        "start": start_date,
        "end": end_date,
        "store": active_store
    }
    
    if customer:
        sql_parts.append("AND LOWER(i.customer_name) LIKE LOWER(:customer)")
        params["customer"] = f"%{customer}%"
        
    if vendor:
        sql_parts.append("AND p.vendor_name = :vendor")
        params["vendor"] = vendor

    query = text(f"""
        SELECT 
            i.invoice_no,
            i.customer_name,
            ii.part_no,
            ii.description,
            ii.quantity,
            ii.rate,
            ii.amount as taxable_amount,
            COALESCE(ii.purchase_rate, 0.0) as purchase_rate,
            i.date
        {chr(10).join(sql_parts)}
    """)

    result = db.execute(query, params).fetchall()
    db.close()

    # Build list of dicts for pandas
    rows = []
    for r in result:
        taxable = float(r.taxable_amount or 0.0)
        gst = taxable * 0.18
        grand = taxable + gst
        cgst = gst / 2
        sgst = gst / 2
        qty = float(r.quantity or 0.0)
        prate = float(r.purchase_rate or 0.0)
        pcost = prate * qty
        profit = taxable - pcost
        
        rows.append({
            "Invoice No": r.invoice_no,
            "Customer": r.customer_name,
            "Part No": r.part_no,
            "Description": r.description,
            "Qty": qty,
            "Selling Rate": r.rate,
            "Taxable Amount": round(taxable, 2),
            "CGST": round(cgst, 2),
            "SGST": round(sgst, 2),
            "Grand Total": round(grand, 2),
            "Purchase Rate": prate,
            "Total Purchase Cost": round(pcost, 2),
            "Profit": round(profit, 2),
            "Date": safe_format_datetime(r.date)
        })

    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        "Invoice No", "Customer", "Part No", "Description", "Qty", "Selling Rate",
        "Taxable Amount", "CGST", "SGST", "Grand Total", "Purchase Rate", "Total Purchase Cost", "Profit", "Date"
    ])

    total_orders = len(df["Invoice No"].unique()) if not df.empty else 0
    total_sales = df["Taxable Amount"].sum() if not df.empty else 0.0
    total_cgst = df["CGST"].sum() if not df.empty else 0.0
    total_sgst = df["SGST"].sum() if not df.empty else 0.0
    grand_total = df["Grand Total"].sum() if not df.empty else 0.0
    total_cost = df["Total Purchase Cost"].sum() if not df.empty else 0.0
    total_profit = df["Profit"].sum() if not df.empty else 0.0

    summary_df = pd.DataFrame({
        "Metric": [
            "Total Orders",
            "Total Sales (Taxable)",
            "Total CGST",
            "Total SGST",
            "Grand Total (incl. GST)",
            "Total Purchase Cost",
            "Total Profit"
        ],
        "Value": [
            total_orders,
            round(total_sales, 2),
            round(total_cgst, 2),
            round(total_sgst, 2),
            round(grand_total, 2),
            round(total_cost, 2),
            round(total_profit, 2)
        ]
    })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Sales Data", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sales.xlsx"}
    )
@app.post("/upload_excel")
def upload_excel(file: UploadFile = File(...)):
    if file.filename.endswith(".csv"):
        df = pd.read_csv(file.file)
    else:
        df = pd.read_excel(file.file, engine="openpyxl")
    df.columns = [col.strip().upper() for col in df.columns]

    db = SessionLocal()
    db.query(OrderItems).delete()

    for _, row in df.iterrows():
        part_no = str(row.get("PART NO", "")).strip()
        if not part_no:
            continue

        db.add(OrderItems(
            part_no=part_no,
            description=row.get("PART DESC", ""),
            mrp=float(str(row.get("MRP", 0)).replace(",", "")),  # ✅ FIX
            hsn=row.get("HSN", "")
        ))

    db.commit()
    db.close()

@app.post("/upload_stock")
def upload_stock(request: Request, file: UploadFile = File(...), mode: str = Form("append")):
    active_store = get_active_store(request)
    db = SessionLocal()
    try:
        # Read Excel or CSV
        if file.filename.endswith(".csv"):
            df = pd.read_csv(file.file)
        else:
            df = pd.read_excel(file.file, engine="openpyxl")

        # Standardize column headers to uppercase and strip whitespace
        df.columns = [col.strip().upper() for col in df.columns]
        df = df.fillna("")

        if mode == "replace":
            db.execute(text("DELETE FROM products WHERE store = :store"), {"store": active_store})
            db.commit()

        imported_count = 0
        for _, row in df.iterrows():
            part_no = str(row.get("PART NO", row.get("PART_NO", row.get("PARTNUMBER", "")))).strip().upper()
            if not part_no:
                continue

            description = str(row.get("DESCRIPTION", row.get("PART DESC", row.get("PART DESCRIPTION", "")))).strip()
            hsn = str(row.get("HSN", row.get("HSN CODE", ""))).strip()
            
            try:
                gst = float(row.get("GST", row.get("GST %", 18.0)))
            except (ValueError, TypeError):
                gst = 18.0

            try:
                qty = int(row.get("QTY", row.get("QUANTITY", row.get("STOCK", 0))))
            except (ValueError, TypeError):
                qty = 0

            try:
                rate = float(row.get("RATE", row.get("PURCHASE RATE", row.get("PRICE", 0.0))))
            except (ValueError, TypeError):
                rate = 0.0

            try:
                discount = float(row.get("DISCOUNT", row.get("DISC", row.get("DISCOUNT %", 0.0))))
            except (ValueError, TypeError):
                discount = 0.0

            vendor_name = str(row.get("VENDOR NAME", row.get("VENDOR", ""))).strip()
            vendor_address = str(row.get("VENDOR ADDRESS", row.get("ADDRESS", ""))).strip()
            vendor_mobile = str(row.get("VENDOR MOBILE", row.get("MOBILE", ""))).strip()
            vendor_gstin = str(row.get("VENDOR GSTIN", row.get("GSTIN", ""))).strip()
            vendor_email = str(row.get("VENDOR EMAIL", row.get("EMAIL", ""))).strip()
            location = str(row.get("LOCATION", row.get("STORE LOCATION", row.get("RACK", row.get("SHELF", row.get("BIN", row.get("PLACE", ""))))))).strip()

            amount = (qty * rate) - ((qty * rate) * (discount / 100)) if qty > 0 else 0.0

            existing = None
            if mode == "append":
                existing = db.query(Product).filter(Product.part_no == part_no, Product.store == active_store).first()

            if existing:
                existing.quantity += qty
                if rate > 0:
                    existing.rate = rate
                if discount > 0:
                    existing.discount = discount
                existing.amount = (existing.quantity * existing.rate) - ((existing.quantity * existing.rate) * (existing.discount / 100))
                if location:
                    existing.location = location
                if vendor_name:
                    existing.vendor_name = vendor_name
                if vendor_address:
                    existing.vendor_address = vendor_address
                if vendor_mobile:
                    existing.vendor_mobile = vendor_mobile
                if vendor_gstin:
                    existing.vendor_gstin = vendor_gstin
                if vendor_email:
                    existing.vendor_email = vendor_email
            else:
                new_prod = Product(
                    part_no=part_no,
                    description=description,
                    hsn=hsn,
                    gst=gst,
                    quantity=qty,
                    rate=rate,
                    discount=discount,
                    amount=round(amount, 2),
                    location=location,
                    store=active_store,
                    vendor_name=vendor_name,
                    vendor_address=vendor_address,
                    vendor_mobile=vendor_mobile,
                    vendor_gstin=vendor_gstin,
                    vendor_email=vendor_email
                )
                db.add(new_prod)
            imported_count += 1

        db.commit()
        return {"message": f"Successfully imported {imported_count} products ({mode} mode)."}

    except Exception as e:
        db.rollback()
        return {"error": str(e)}, 500
    finally:
        db.close()


@app.post("/update_location/{product_id}")
async def update_product_location(product_id: int, request: Request):
    data = await request.json()
    new_location = str(data.get("location", "")).strip()
    db = SessionLocal()
    try:
        product = db.query(Product).filter(Product.id == product_id).first()
        if not product:
            return {"error": "Product not found"}
        product.location = new_location
        db.commit()
        return {"ok": True, "location": new_location}
    except Exception as e:
        db.rollback()
        return {"error": str(e)}
    finally:
        db.close()


@app.get("/export_stock_excel")
def export_stock_excel(request: Request, q: str = Query("")):
    active_store = get_active_store(request)
    db = SessionLocal()
    try:
        from sqlalchemy import func
        q_clean = q.strip().upper()
        query = db.query(Product).filter(Product.store == active_store)
        if q_clean:
            query = query.filter(
                (func.upper(Product.part_no).like(f"%{q_clean}%")) |
                (func.upper(Product.description).like(f"%{q_clean}%")) |
                (func.upper(Product.location).like(f"%{q_clean}%"))
            )
        
        products = query.order_by(Product.part_no.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Inventory"

        # Enable grid lines
        ws.views.sheetView[0].showGridLines = True

        # Header style
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = [
            "Part No", "Description", "HSN", "GST %", 
            "Quantity", "Stock Status", "Purchase Rate (₹)", 
            "Discount %", "Taxable Amount (₹)", "Store Location", "Vendor Name"
        ]
        
        ws.append(headers)
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        total_qty = 0
        total_value = 0.0
        in_stock_count = 0
        out_stock_count = 0

        data_row_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        for p in products:
            qty = int(p.quantity or 0)
            rate = float(p.rate or 0.0)
            discount = float(p.discount or 0.0)
            amt = float(p.amount or 0.0)
            if qty <= 0:
                amt = 0.0
                status = "Out of Stock"
                out_stock_count += 1
            elif qty < 5:
                status = "Low Stock"
                in_stock_count += 1
                total_qty += qty
                total_value += amt
            else:
                status = "In Stock"
                in_stock_count += 1
                total_qty += qty
                total_value += amt

            row_data = [
                p.part_no or "",
                p.description or "",
                p.hsn or "",
                f"{p.gst or 18}%",
                qty,
                status,
                round(rate, 2),
                f"{discount}%",
                round(amt, 2),
                p.location or "Unassigned",
                p.vendor_name or ""
            ]
            ws.append(row_data)

            r_idx = ws.max_row
            ws.cell(row=r_idx, column=1).alignment = Alignment(horizontal="left")
            ws.cell(row=r_idx, column=2).alignment = Alignment(horizontal="left")
            ws.cell(row=r_idx, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=r_idx, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=r_idx, column=5).alignment = Alignment(horizontal="right")
            ws.cell(row=r_idx, column=6).alignment = Alignment(horizontal="center")
            ws.cell(row=r_idx, column=7).alignment = Alignment(horizontal="right")
            ws.cell(row=r_idx, column=8).alignment = Alignment(horizontal="center")
            ws.cell(row=r_idx, column=9).alignment = Alignment(horizontal="right")
            ws.cell(row=r_idx, column=10).alignment = Alignment(horizontal="center")
            ws.cell(row=r_idx, column=11).alignment = Alignment(horizontal="left")

            for col_idx in range(1, 12):
                ws.cell(row=r_idx, column=col_idx).border = data_row_border

        ws.append([])

        summary_title_font = Font(name="Calibri", size=11, bold=True)
        ws.append(["Stock Summary"])
        ws.cell(row=ws.max_row, column=1).font = summary_title_font

        ws.append(["Total Unique Products", len(products)])
        ws.append(["In Stock Items", in_stock_count])
        ws.append(["Out of Stock Items", out_stock_count])
        ws.append(["Total Stock Quantity", total_qty])
        ws.append(["Total Inventory Valuation (₹)", round(total_value, 2)])

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "stock_inventory.xlsx" if not q_clean else f"stock_inventory_{q_clean}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    finally:
        db.close()



# ---------------- ADD PRODUCT ----------------
@app.post("/add_stock/{id}")
def add_stock(id: int, qty: int = Form(...)):
    db = SessionLocal()

    product = db.query(Product).filter(Product.id == id).first()

    if not product:
        db.close()
        return {"error": "Product not found"}

    # ✅ Add stock
    product.quantity += qty

    if product.quantity <= 0:
        product.amount = 0
    else:
        product.amount = (product.quantity * (product.rate or 0)) - (
            (product.quantity * (product.rate or 0)) * ((product.discount or 0) / 100)
        )

    db.commit()
    db.close()

    return RedirectResponse("/", status_code=303)

@app.post("/issue/{id}")
def issue_stock(id: int, qty: int = Form(...)):
    db = SessionLocal()

    product = db.query(Product).filter(Product.id == id).first()

    if not product:
        db.close()
        return {"error": "Product not found"}

    if product.quantity < qty:
        db.close()
        return {"error": "Not enough stock"}

    # ✅ STEP 1: Reduce stock
    product.quantity -= qty

    # ✅ STEP 2: Recalculate amount
    if product.quantity <= 0:
        product.quantity = 0
        product.amount = 0
    else:
        product.amount = (product.quantity * (product.rate or 0)) - (
            (product.quantity * (product.rate or 0)) * ((product.discount or 0) / 100)
        )

    # ✅ STEP 3: Save changes
    db.commit()
    db.refresh(product)
    db.close()

    return RedirectResponse("/", status_code=303)
@app.post("/add")
def add_product(
    request: Request,
    part_no: str = Form(...),
    description: str = Form(...),
    hsn: str = Form(...),
    gst: float = Form(...),
    quantity: int = Form(...),
    rate: float = Form(...),
    discount: float = Form(0),
    location: str = Form(None),
    vendor_name: str = Form(None),
    vendor_address: str = Form(None),
    vendor_mobile: str = Form(None),
    vendor_gstin: str = Form(None),
    vendor_email: str = Form(None)
):
    active_store = get_active_store(request)
    db = SessionLocal()

    # Rate fallback if not provided or 0
    if rate <= 0:
        price = db.query(OrderItems).filter(OrderItems.part_no == part_no, OrderItems.store == active_store).first()
        if price and price.mrp > 0:
            rate = price.mrp

    # 1. Save or Update Vendor details if vendor_name is provided
    if vendor_name and vendor_name.strip():
        v_name = vendor_name.strip()
        existing_vendor = db.query(Vendor).filter(Vendor.name == v_name, Vendor.store == active_store).first()
        if existing_vendor:
            if vendor_address is not None:
                existing_vendor.address = vendor_address.strip()
            if vendor_mobile is not None:
                existing_vendor.mobile_num = vendor_mobile.strip()
            if vendor_gstin is not None:
                existing_vendor.gstin = vendor_gstin.strip()
            if vendor_email is not None:
                existing_vendor.email_id = vendor_email.strip()
        else:
            new_vendor = Vendor(
                name=v_name,
                address=vendor_address.strip() if vendor_address else "",
                mobile_num=vendor_mobile.strip() if vendor_mobile else "",
                gstin=vendor_gstin.strip() if vendor_gstin else "",
                email_id=vendor_email.strip() if vendor_email else "",
                store=active_store
            )
            db.add(new_vendor)
        db.commit()

    # 2. Record Purchase if quantity > 0
    if quantity > 0:
        purchase_amt = (quantity * rate) - ((quantity * rate) * (discount / 100))
        new_purchase = Purchase(
            vendor_name=vendor_name.strip() if (vendor_name and vendor_name.strip()) else "Local Vendor",
            part_no=part_no.strip(),
            description=description.strip(),
            hsn=hsn.strip(),
            quantity=quantity,
            rate=rate,
            discount=discount,
            amount=round(purchase_amt, 2),
            date=datetime.now(),
            store=active_store
        )
        db.add(new_purchase)

    # 3. Update stock (existing logic)
    existing = db.query(Product).filter(Product.part_no == part_no, Product.store == active_store).first()

    if existing:
        existing.quantity += quantity

        # update rate if valid
        if rate > 0:
            existing.rate = rate

        # calculate amount
        if existing.quantity <= 0:
            existing.quantity = 0
            existing.amount = 0
        else:
            existing.amount = (existing.quantity * existing.rate) - (
                (existing.quantity * existing.rate) * (existing.discount / 100)
            )

        # Update vendor details and location
        if location and location.strip():
            existing.location = location.strip()
        if vendor_name and vendor_name.strip():
            existing.vendor_name = vendor_name.strip()
            existing.vendor_address = vendor_address.strip() if vendor_address else ""
            existing.vendor_mobile = vendor_mobile.strip() if vendor_mobile else ""
            existing.vendor_gstin = vendor_gstin.strip() if vendor_gstin else ""
            existing.vendor_email = vendor_email.strip() if vendor_email else ""

    else:
        # new product
        if quantity <= 0:
            amount = 0
        else:
            amount = (quantity * rate) - (
                (quantity * rate) * (discount / 100)
            )

        new_product = Product(
            part_no=part_no,
            description=description,
            hsn=hsn,
            gst=gst,
            quantity=quantity,
            rate=rate,
            discount=discount,
            amount=amount,
            location=location.strip() if (location and location.strip()) else "",
            store=active_store,
            vendor_name=vendor_name.strip() if (vendor_name and vendor_name.strip()) else "",
            vendor_address=vendor_address.strip() if vendor_address else "",
            vendor_mobile=vendor_mobile.strip() if vendor_mobile else "",
            vendor_gstin=vendor_gstin.strip() if vendor_gstin else "",
            vendor_email=vendor_email.strip() if vendor_email else ""
        )

        db.add(new_product)

    # SINGLE COMMIT
    db.commit()
    db.close()

    return RedirectResponse("/", status_code=303)


@app.post("/add_purchase_bill")
async def add_purchase_bill(request: Request):
    active_store = get_active_store(request)
    data = await request.json()
    db = SessionLocal()
    
    vendor_name = data.get("vendor_name", "").strip()
    vendor_address = data.get("vendor_address", "").strip()
    vendor_mobile = data.get("vendor_mobile", "").strip()
    vendor_gstin = data.get("vendor_gstin", "").strip()
    vendor_email = data.get("vendor_email", "").strip()
    items = data.get("items", [])
    
    if not items:
        db.close()
        return {"error": "No items in purchase bill"}
        
    # 1. Save or Update Vendor Details
    if vendor_name:
        existing_vendor = db.query(Vendor).filter(Vendor.name == vendor_name, Vendor.store == active_store).first()
        if existing_vendor:
            existing_vendor.address = vendor_address
            existing_vendor.mobile_num = vendor_mobile
            existing_vendor.gstin = vendor_gstin
            existing_vendor.email_id = vendor_email
        else:
            new_vendor = Vendor(
                name=vendor_name,
                address=vendor_address,
                mobile_num=vendor_mobile,
                gstin=vendor_gstin,
                email_id=vendor_email,
                store=active_store
            )
            db.add(new_vendor)
        db.commit()
        
    # 2. Process items
    for item in items:
        part_no = item.get("part_no", "").strip()
        description = item.get("description", "").strip()
        hsn = item.get("hsn", "").strip()
        gst = float(item.get("gst", 18.0) or 18.0)
        qty = int(item.get("qty", 0) or 0)
        rate = float(item.get("rate", 0.0) or 0.0)
        discount = float(item.get("discount", 0.0) or 0.0)
        
        if not part_no:
            continue
            
        # Rate fallback if not provided or 0
        if rate <= 0:
            price = db.query(OrderItems).filter(OrderItems.part_no == part_no, OrderItems.store == active_store).first()
            if price and price.mrp > 0:
                rate = price.mrp
                
        # Record Purchase if qty > 0
        if qty > 0:
            purchase_amt = (qty * rate) - ((qty * rate) * (discount / 100))
            new_purchase = Purchase(
                vendor_name=vendor_name if vendor_name else "Local Vendor",
                part_no=part_no,
                description=description,
                hsn=hsn,
                quantity=qty,
                rate=rate,
                discount=discount,
                amount=round(purchase_amt, 2),
                date=datetime.now(),
                store=active_store
            )
            db.add(new_purchase)
            
        # Update or Create Product Stock
        existing = db.query(Product).filter(Product.part_no == part_no, Product.store == active_store).first()
        item_loc = str(item.get("location", "") or "").strip()
        if existing:
            existing.quantity += qty
            if rate > 0:
                existing.rate = rate
            if description:
                existing.description = description
            if hsn:
                existing.hsn = hsn
            existing.gst = gst
            existing.discount = discount
            if item_loc:
                existing.location = item_loc
            
            # Recalculate amount
            if existing.quantity <= 0:
                existing.quantity = 0
                existing.amount = 0
            else:
                existing.amount = (existing.quantity * existing.rate) - (
                    (existing.quantity * existing.rate) * (existing.discount / 100)
                )
                
            # Update vendor details on product
            if vendor_name:
                existing.vendor_name = vendor_name
                existing.vendor_address = vendor_address
                existing.vendor_mobile = vendor_mobile
                existing.vendor_gstin = vendor_gstin
                existing.vendor_email = vendor_email
        else:
            # new product
            if qty <= 0:
                amount = 0
            else:
                amount = (qty * rate) - ((qty * rate) * (discount / 100))
                
            new_product = Product(
                part_no=part_no,
                description=description,
                hsn=hsn,
                gst=gst,
                quantity=qty,
                rate=rate,
                discount=discount,
                amount=round(amount, 2),
                location=item_loc,
                store=active_store,
                vendor_name=vendor_name,
                vendor_address=vendor_address,
                vendor_mobile=vendor_mobile,
                vendor_gstin=vendor_gstin,
                vendor_email=vendor_email
            )
            db.add(new_product)
            
    db.commit()
    db.close()
    return {"ok": True}

# ---------------- GET PRICE ----------------

@app.get("/get_price/{part_no}")
def get_price(request: Request, part_no: str):
    active_store = get_active_store(request)
    db = SessionLocal()

    # Step 1: Check Stock Data (products table) FIRST
    prod = db.query(Product).filter(Product.part_no == part_no, Product.store == active_store).first()

    # Step 2: Check Price Master (order_items table) as fallback
    price = db.query(OrderItems).filter(
        OrderItems.part_no == part_no,
        OrderItems.store == active_store
    ).first()

    db.close()

    if prod and prod.rate and float(prod.rate) > 0:
        rate = float(prod.rate)
        description = prod.description or (price.description if price else "")
        hsn = prod.hsn or (price.hsn if price else "")
    else:
        rate = float(price.mrp) if (price and price.mrp) else (float(prod.rate) if (prod and prod.rate) else 0.0)
        description = price.description if price else (prod.description if prod else "")
        hsn = price.hsn if price else (prod.hsn if prod else "")

    return {
        "rate": rate,
        "hsn": hsn,
        "description": description,
        "location": prod.location if (prod and prod.location) else "",
        "vendor_name": prod.vendor_name if prod else "",
        "vendor_address": prod.vendor_address if prod else "",
        "vendor_mobile": prod.vendor_mobile if prod else "",
        "vendor_gstin": prod.vendor_gstin if prod else "",
        "vendor_email": prod.vendor_email if prod else ""
    }


@app.get("/upload_page", response_class=HTMLResponse)
def upload_page():
    return """
    <html>
    <body>
        <h2>Upload Order Excel</h2>
        <form action="/upload_order" method="post" enctype="multipart/form-data">
            <input type="file" name="file" />
            <button type="submit">Upload</button>
        </form>
    </body>
    </html>
    """

@app.post("/process_order")
def process_order(file: UploadFile = File(...)):
    df = pd.read_excel(file.file)
    df.columns = [col.strip().upper() for col in df.columns]


    db = SessionLocal()
    result = []

    for _, row in df.iterrows():
        part_no = str(row.get("PART NO", "")).strip()
        qty = int(row.get("QTY", 0))

        price = db.query(OrderItems).filter(OrderItems.part_no == part_no).first()
        stock = db.query(Product).filter(Product.part_no == part_no).first()

        result.append({
            "part_no": part_no,
            "description": price.description if price else "",
            "qty": qty,
            "rate": price.mrp if price else 0,
            "available": stock.quantity if stock else 0,
            "discount": 0
        })

    db.close()
    return result
@app.get("/get_rate/{part_no}")
def get_rate(request: Request, part_no: str):
    active_store = get_active_store(request)
    db = SessionLocal()

    # Step 1: Check products table (Stock Data) FIRST
    product = db.query(Product).filter(Product.part_no == part_no, Product.store == active_store).first()
    stock = int(product.quantity or 0) if product else 0

    # Step 2: Check order_items (price master) as fallback
    item = db.execute(
        text("SELECT mrp, description, hsn FROM order_items WHERE part_no = :p AND store = :store"),
        {"p": part_no, "store": active_store}
    ).fetchone()

    if product and product.rate and float(product.rate) > 0:
        rate = float(product.rate)
        description = product.description or (item.description if item else "")
        hsn = product.hsn or (str(item.hsn) if item else "")
    else:
        rate = float(item.mrp) if (item and item.mrp) else (float(product.rate) if (product and product.rate) else 0.0)
        description = item.description if item else (product.description if product else "")
        hsn = str(item.hsn) if item else (product.hsn if product else "")

    db.close()
    return {"rate": rate, "description": description, "hsn": hsn, "stock": stock}


# ---------------- PRICE MASTER (orders.csv) ----------------

@app.get("/reimport_orders")
def reimport_orders(request: Request):
    """Admin: reload all products from orders.csv into order_items"""
    active_store = get_active_store(request)
    db = SessionLocal()
    try:
        csv_path = os.path.join(os.path.dirname(__file__), "orders.csv")
        if not os.path.exists(csv_path):
            return {"error": "orders.csv not found"}
        df = pd.read_csv(csv_path, dtype=str, on_bad_lines="skip")
        df.columns = [col.strip() for col in df.columns]
        db.execute(text("DELETE FROM order_items WHERE store = :store"), {"store": active_store})
        batch, inserted = [], 0
        for _, row in df.iterrows():
            part_no = str(row.get("Part No", "") or "").strip()
            if not part_no:
                continue
            batch.append({
                "pn":   part_no,
                "desc": str(row.get("Part Desc", "") or "").strip(),
                "hsn":  str(row.get("HSN", "") or "").strip(),
                "mrp":  float(row.get("MRP", 0) or 0)
            })
            if len(batch) >= 500:
                params = {"store": active_store}
                values_clauses = []
                for i, item in enumerate(batch):
                    values_clauses.append(f"(:pn_{i}, :desc_{i}, :hsn_{i}, :mrp_{i}, :store)")
                    params[f"pn_{i}"] = item["pn"]
                    params[f"desc_{i}"] = item["desc"]
                    params[f"hsn_{i}"] = item["hsn"]
                    params[f"mrp_{i}"] = item["mrp"]
                sql = f"INSERT INTO order_items (part_no, description, hsn, mrp, store) VALUES {', '.join(values_clauses)}"
                db.execute(text(sql), params)
                db.commit()
                inserted += len(batch)
                batch = []
        if batch:
            params = {"store": active_store}
            values_clauses = []
            for i, item in enumerate(batch):
                values_clauses.append(f"(:pn_{i}, :desc_{i}, :hsn_{i}, :mrp_{i}, :store)")
                params[f"pn_{i}"] = item["pn"]
                params[f"desc_{i}"] = item["desc"]
                params[f"hsn_{i}"] = item["hsn"]
                params[f"mrp_{i}"] = item["mrp"]
            sql = f"INSERT INTO order_items (part_no, description, hsn, mrp, store) VALUES {', '.join(values_clauses)}"
            db.execute(text(sql), params)
            db.commit()
            inserted += len(batch)
        return {"ok": True, "imported": inserted}
    except Exception as e:
        return {"error": str(e)}
    finally:
        db.close()

@app.get("/search_parts")
def search_parts(request: Request, q: str = ""):
    """Autocomplete: search order_items by part_no or description"""
    active_store = get_active_store(request)
    db = SessionLocal()
    q = q.strip().upper()
    results = db.execute(text("""
        SELECT part_no, description, mrp, hsn
        FROM order_items
        WHERE store = :store AND (UPPER(part_no) LIKE :q OR UPPER(description) LIKE :q)
        LIMIT 20
    """), {"q": f"%{q}%", "store": active_store}).fetchall()
    db.close()
    return [{"part_no": r.part_no, "description": r.description,
             "rate": float(r.mrp or 0), "hsn": str(r.hsn or "")} for r in results]

@app.get("/search_vendors")
def search_vendors(request: Request, q: str = ""):
    """Autocomplete: search vendors by name"""
    active_store = get_active_store(request)
    db = SessionLocal()
    q_clean = f"%{q.strip().upper()}%"
    rows = db.execute(text("""
        SELECT name, address, mobile_num, gstin, email_id
        FROM vendors
        WHERE store = :store AND UPPER(name) LIKE :q
        LIMIT 10
    """), {"q": q_clean, "store": active_store}).fetchall()
    db.close()
    return [{
        "name": r.name,
        "address": r.address or "",
        "mobile_num": r.mobile_num or "",
        "gstin": r.gstin or "",
        "email_id": r.email_id or ""
    } for r in rows]

@app.get("/price_master")
def price_master_list(request: Request, page: int = 1, q: str = ""):
    """List all products in order_items (price master) with pagination"""
    active_store = get_active_store(request)
    db = SessionLocal()
    per_page = 100
    offset = (page - 1) * per_page
    q_clean = f"%{q.strip().upper()}%"
    rows = db.execute(text("""
        SELECT id, part_no, description, hsn, mrp
        FROM order_items
        WHERE store = :store AND (UPPER(part_no) LIKE :q OR UPPER(description) LIKE :q)
        ORDER BY part_no
        LIMIT :lim OFFSET :off
    """), {"q": q_clean, "lim": per_page, "off": offset, "store": active_store}).fetchall()
    total = db.execute(text("""
        SELECT COUNT(*) FROM order_items
        WHERE store = :store AND (UPPER(part_no) LIKE :q OR UPPER(description) LIKE :q)
    """), {"q": q_clean, "store": active_store}).scalar()
    db.close()
    return {"total": total, "page": page, "per_page": per_page,
            "items": [{"id": r.id, "part_no": r.part_no, "description": r.description,
                       "hsn": r.hsn, "mrp": float(r.mrp or 0)} for r in rows]}

@app.post("/price_master/add")
async def price_master_add(request: Request):
    active_store = get_active_store(request)
    data = await request.json()
    db = SessionLocal()
    existing = db.query(OrderItems).filter(OrderItems.part_no == data["part_no"].strip(), OrderItems.store == active_store).first()
    if existing:
        db.close()
        return {"error": "Part No already exists in this store"}
    item = OrderItems(
        part_no=data["part_no"].strip(),
        description=data.get("description", "").strip(),
        hsn=data.get("hsn", "").strip(),
        mrp=float(data.get("mrp", 0)),
        store=active_store
    )
    db.add(item)
    db.commit()
    db.close()
    return {"ok": True}

@app.post("/price_master/update/{item_id}")
async def price_master_update(item_id: int, request: Request):
    data = await request.json()
    db = SessionLocal()
    item = db.query(OrderItems).filter(OrderItems.id == item_id).first()
    if item:
        item.description = data.get("description", item.description)
        item.hsn = data.get("hsn", item.hsn)
        item.mrp = float(data.get("mrp", item.mrp))
        db.commit()
    db.close()
    return {"ok": True}

@app.post("/price_master/delete/{item_id}")
def price_master_delete(item_id: int):
    db = SessionLocal()
    db.query(OrderItems).filter(OrderItems.id == item_id).delete()
    db.commit()
    db.close()
    return {"ok": True}

# ---------------- QUOTATION (no stock change) ----------------

def generate_quotation_no(db):
    import calendar
    today = datetime.now()
    year_month = today.strftime("%Y-%m")
    first_day = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_day_num = calendar.monthrange(today.year, today.month)[1]
    last_day = today.replace(day=last_day_num, hour=23, minute=59, second=59, microsecond=999999)
    count = db.execute(text("""
        SELECT COUNT(*) FROM quotations WHERE date >= :fd AND date <= :ld
    """), {"fd": first_day, "ld": last_day}).scalar()
    return f"QT-{year_month}-{str(count+1).zfill(3)}"

@app.post("/save_quotation")
async def save_quotation(request: Request):
    """Save quotation to DB — NO stock change, NO sales report impact"""
    active_store = get_active_store(request)
    data = await request.json()
    db = SessionLocal()
    rows = data.get("rows", [])
    customer_name = data.get("customer_name", "")
    customer_address = data.get("customer_address", "")
    customer_gstin = data.get("customer_gstin", "")
    customer_mobile = data.get("customer_mobile", "")
    customer_email = data.get("customer_email", "")

    if not rows:
        return {"error": "No products"}

    # 1. Save or Update Vendor details if customer_name is provided
    if customer_name and customer_name.strip():
        c_name = customer_name.strip()
        existing_vendor = db.query(Vendor).filter(Vendor.name == c_name, Vendor.store == active_store).first()
        if existing_vendor:
            if customer_address: existing_vendor.address = customer_address.strip()
            if customer_mobile: existing_vendor.mobile_num = customer_mobile.strip()
            if customer_gstin: existing_vendor.gstin = customer_gstin.strip()
            if customer_email: existing_vendor.email_id = customer_email.strip()
        else:
            new_vendor = Vendor(
                name=c_name,
                address=customer_address.strip() if customer_address else "",
                mobile_num=customer_mobile.strip() if customer_mobile else "",
                gstin=customer_gstin.strip() if customer_gstin else "",
                email_id=customer_email.strip() if customer_email else "",
                store=active_store
            )
            db.add(new_vendor)
        db.commit()

    total_amount = sum(r["rate"] * r["qty"] * (1 - r.get("discount", 0)/100) for r in rows)
    grand_total = total_amount * 1.18
    q_no = generate_quotation_no(db)
    quot = Quotation(
        quotation_no=q_no,
        customer_name=customer_name,
        customer_address=customer_address,
        customer_gstin=customer_gstin,
        customer_mobile=customer_mobile,
        customer_email=customer_email,
        date=datetime.now(),
        total_amount=round(total_amount, 2),
        grand_total=round(grand_total, 2),
        store=active_store
    )
    db.add(quot)
    db.flush()
    for r in rows:
        sub = r["rate"] * r["qty"] * (1 - r.get("discount", 0)/100)
        db.add(QuotationItem(
            quotation_id=quot.id,
            part_no=r.get("part_no", ""),
            description=r.get("description", ""),
            rate=r["rate"],
            qty=r["qty"],
            discount=r.get("discount", 0),
            amount=round(sub, 2),
            hsn=r.get("hsn", "")
        ))
    db.commit()
    q_id = quot.id
    db.close()
    return {"ok": True, "quotation_no": q_no, "id": q_id}

@app.get("/quotations")
def list_quotations(request: Request):
    active_store = get_active_store(request)
    db = SessionLocal()
    rows = db.execute(text("""
        SELECT id, quotation_no, customer_name, customer_address, customer_gstin, customer_mobile, customer_email, date, grand_total
        FROM quotations WHERE store = :store ORDER BY id DESC LIMIT 50
    """), {"store": active_store}).fetchall()
    db.close()
    return [{"id": r.id, "quotation_no": r.quotation_no,
             "customer_name": r.customer_name,
             "customer_address": r.customer_address,
             "customer_gstin": r.customer_gstin,
             "customer_mobile": r.customer_mobile,
             "customer_email": r.customer_email,
             "date": str(r.date)[:10],
             "grand_total": float(r.grand_total or 0)} for r in rows]

@app.get("/quotation_details/{q_id}")
def quotation_details(q_id: int, request: Request):
    active_store = get_active_store(request)
    db = SessionLocal()
    quot = db.query(Quotation).filter(Quotation.id == q_id, Quotation.store == active_store).first()
    if not quot:
        db.close()
        return {"error": "Quotation not found"}
    items = db.query(QuotationItem).filter(QuotationItem.quotation_id == q_id).all()
    db.close()
    return {
        "id": quot.id,
        "quotation_no": quot.quotation_no,
        "customer_name": quot.customer_name or "",
        "customer_address": quot.customer_address or "",
        "customer_gstin": quot.customer_gstin or "",
        "customer_mobile": quot.customer_mobile or "",
        "customer_email": quot.customer_email or "",
        "total_amount": float(quot.total_amount or 0.0),
        "grand_total": float(quot.grand_total or 0.0),
        "items": [{
            "part_no": it.part_no,
            "description": it.description,
            "rate": float(it.rate or 0.0),
            "qty": int(it.qty or 0),
            "discount": float(it.discount or 0.0),
            "amount": float(it.amount or 0.0),
            "hsn": it.hsn
        } for it in items]
    }

@app.post("/update_quotation/{q_id}")
async def update_quotation(q_id: int, request: Request):
    active_store = get_active_store(request)
    data = await request.json()
    db = SessionLocal()
    quot = db.query(Quotation).filter(Quotation.id == q_id, Quotation.store == active_store).first()
    if not quot:
        db.close()
        return {"error": "Quotation not found"}

    rows = data.get("rows", [])
    if not rows:
        db.close()
        return {"error": "No products provided"}

    quot.customer_name = data.get("customer_name", quot.customer_name)
    quot.customer_address = data.get("customer_address", quot.customer_address)
    quot.customer_gstin = data.get("customer_gstin", quot.customer_gstin)
    quot.customer_mobile = data.get("customer_mobile", quot.customer_mobile)
    quot.customer_email = data.get("customer_email", quot.customer_email)

    db.query(QuotationItem).filter(QuotationItem.quotation_id == q_id).delete()

    total_amount = sum(r["rate"] * r["qty"] * (1 - r.get("discount", 0)/100) for r in rows)
    grand_total = total_amount * 1.18

    quot.total_amount = round(total_amount, 2)
    quot.grand_total = round(grand_total, 2)

    for r in rows:
        sub = r["rate"] * r["qty"] * (1 - r.get("discount", 0)/100)
        db.add(QuotationItem(
            quotation_id=quot.id,
            part_no=r.get("part_no", ""),
            description=r.get("description", ""),
            rate=float(r["rate"]),
            qty=int(r["qty"]),
            discount=float(r.get("discount", 0)),
            amount=round(sub, 2),
            hsn=r.get("hsn", "")
        ))
    q_no = quot.quotation_no
    q_id_val = quot.id
    db.commit()
    db.close()
    return {"ok": True, "quotation_no": q_no, "id": q_id_val}

# ---------------- TAX INVOICE EDITING & CORRECTION ----------------

@app.get("/invoices_list")
def list_invoices(request: Request, q: str = ""):
    active_store = get_active_store(request)
    db = SessionLocal()
    q_clean = f"%{q.strip().upper()}%"
    sql = """
        SELECT id, invoice_no, customer_name, customer_address, customer_gstin, customer_mobile, customer_email, date, total_amount, gst_amount, grand_total
        FROM invoices
        WHERE store = :store AND (UPPER(invoice_no) LIKE :q OR UPPER(customer_name) LIKE :q)
        ORDER BY date DESC
        LIMIT 50
    """
    rows = db.execute(text(sql), {"store": active_store, "q": q_clean}).fetchall()
    db.close()
    return [{
        "id": r.id,
        "invoice_no": r.invoice_no,
        "customer_name": r.customer_name,
        "customer_address": r.customer_address,
        "customer_gstin": r.customer_gstin,
        "customer_mobile": r.customer_mobile,
        "customer_email": r.customer_email,
        "date": safe_format_datetime(r.date),
        "total_amount": float(r.total_amount or 0.0),
        "gst_amount": float(r.gst_amount or 0.0),
        "grand_total": float(r.grand_total or 0.0)
    } for r in rows]

@app.get("/invoice_details/{inv_id}")
def invoice_details(inv_id: int, request: Request):
    active_store = get_active_store(request)
    db = SessionLocal()
    inv = db.query(Invoice).filter(Invoice.id == inv_id, Invoice.store == active_store).first()
    if not inv:
        db.close()
        return {"error": "Invoice not found"}
    items_rows = db.execute(text("""
        SELECT part_no, description, quantity, rate, amount, hsn, purchase_rate
        FROM invoice_items WHERE invoice_id = :iid
    """), {"iid": inv_id}).fetchall()
    db.close()
    items = []
    for r in items_rows:
        sub = r.rate * r.quantity
        taxable = r.amount
        disc_amt = max(0.0, sub - taxable)
        disc_pct = round((disc_amt / sub) * 100, 2) if sub > 0 else 0.0
        items.append({
            "part_no": r.part_no,
            "description": r.description,
            "hsn": r.hsn,
            "rate": float(r.rate),
            "qty": float(r.quantity),
            "discount": disc_pct,
            "taxable": round(taxable, 2),
            "purchase_rate": float(r.purchase_rate or 0.0)
        })
    return {
        "id": inv.id,
        "invoice_no": inv.invoice_no,
        "customer_name": inv.customer_name or "",
        "customer_address": inv.customer_address or "",
        "customer_gstin": inv.customer_gstin or "",
        "customer_mobile": inv.customer_mobile or "",
        "customer_email": inv.customer_email or "",
        "date": safe_format_datetime(inv.date),
        "total_amount": float(inv.total_amount or 0.0),
        "gst_amount": float(inv.gst_amount or 0.0),
        "grand_total": float(inv.grand_total or 0.0),
        "items": items
    }

@app.post("/update_invoice/{inv_id}")
async def update_invoice(inv_id: int, request: Request):
    active_store = get_active_store(request)
    data = await request.json()
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == inv_id, Invoice.store == active_store).first()
        if not inv:
            db.close()
            return {"error": "Invoice not found"}

        rows = data.get("rows", [])
        if not rows:
            db.close()
            return {"error": "No items provided for invoice update"}

        # 1. Revert previous stock subtractions
        old_items = db.execute(text("""
            SELECT part_no, quantity FROM invoice_items WHERE invoice_id = :iid
        """), {"iid": inv_id}).fetchall()

        for old_it in old_items:
            p = db.query(Product).filter(Product.part_no == old_it.part_no, Product.store == active_store).first()
            if p:
                p.quantity += int(old_it.quantity or 0)
                p.amount = (p.quantity * (p.rate or 0)) - ((p.quantity * (p.rate or 0)) * ((p.discount or 0) / 100))

        # 2. Validate and apply new stock quantities
        from collections import defaultdict
        qty_map = defaultdict(float)
        for r in rows:
            qty_map[r["part_no"]] += float(r["qty"])

        for part_no, total_qty in qty_map.items():
            product = db.query(Product).filter(Product.part_no == part_no, Product.store == active_store).first()
            if not product or product.quantity < total_qty:
                raise ValueError(f"Not enough stock for {part_no} (Available: {product.quantity if product else 0}, Required: {total_qty})")

        # Update customer details
        inv.customer_name = data.get("customer_name", inv.customer_name)
        inv.customer_address = data.get("customer_address", inv.customer_address)
        inv.customer_gstin = data.get("customer_gstin", inv.customer_gstin)
        inv.customer_mobile = data.get("customer_mobile", inv.customer_mobile)
        inv.customer_email = data.get("customer_email", inv.customer_email)

        # Clear existing invoice items
        db.execute(text("DELETE FROM invoice_items WHERE invoice_id = :iid"), {"iid": inv_id})

        items = []
        subtotal = 0.0
        for row in rows:
            product = db.query(Product).filter(Product.part_no == row["part_no"], Product.store == active_store).first()
            db_item = db.execute(
                text("SELECT description, mrp, hsn FROM order_items WHERE part_no=:p AND store=:store"),
                {"p": row["part_no"], "store": active_store}
            ).fetchone()

            desc = row.get("description", "") or (product.description if product else "") or (db_item.description if db_item else "")
            hsn = row.get("hsn", "") or (product.hsn if product else "") or (str(db_item.hsn) if db_item else "")

            rate = float(row.get("rate", 0))
            if rate == 0 and product and product.rate and float(product.rate) > 0:
                rate = float(product.rate)
            if rate == 0 and db_item and db_item.mrp and float(db_item.mrp) > 0:
                rate = float(db_item.mrp)

            qty = float(row.get("qty", 1))
            disc = float(row.get("discount", 0))

            purchase_rate = 0.0
            if product:
                purchase_rate = float(product.rate or 0.0)
                product.quantity -= int(qty)
                if product.quantity <= 0:
                    product.quantity = 0
                    product.amount = 0
                else:
                    product.amount = (product.quantity * (product.rate or 0)) - ((product.quantity * (product.rate or 0)) * ((product.discount or 0) / 100))

            base = qty * rate
            discount_amt = base * (disc / 100)
            taxable = base - discount_amt
            subtotal += taxable

            db.execute(text("""
                INSERT INTO invoice_items 
                (invoice_id, part_no, description, quantity, rate, amount, hsn, purchase_rate)
                VALUES (:iid, :p, :d, :q, :r, :amt, :hsn, :prate)
            """), {
                "iid": inv_id,
                "p": row["part_no"],
                "d": desc,
                "q": qty,
                "r": rate,
                "amt": round(taxable, 2),
                "hsn": hsn,
                "prate": purchase_rate
            })

        cgst = round(subtotal * 0.09, 2)
        sgst = round(subtotal * 0.09, 2)
        total = round(subtotal + cgst + sgst, 2)

        inv.total_amount = round(subtotal, 2)
        inv.gst_amount = cgst + sgst
        inv.grand_total = total

        inv_no = inv.invoice_no
        inv_id_val = inv.id
        db.commit()
        db.close()
        return {"ok": True, "invoice_no": inv_no, "id": inv_id_val}

    except Exception as e:
        db.rollback()
        db.close()
        return {"error": str(e)}

@app.post("/download_quotation_pdf")
async def download_quotation_pdf(request: Request):
    """Generate proforma invoice PDF HTML — NO stock change"""
    data = await request.json()
    rows = data.get("rows", [])
    customer_name = data.get("customer_name", "")
    customer_address = data.get("customer_address", "")
    customer_gstin = data.get("customer_gstin", "")
    customer_mobile = data.get("customer_mobile", "")
    customer_email = data.get("customer_email", "")

    items, subtotal = [], 0
    savings = 0.0
    for r in rows:
        sub = r["rate"] * r["qty"]
        disc = sub * (r.get("discount", 0) / 100)
        after = sub - disc
        items.append({
            "part_no": r.get("part_no", ""),
            "description": r.get("description", ""),
            "hsn": r.get("hsn", ""),
            "rate": r["rate"],
            "qty": r["qty"],
            "discount": r.get("discount", 0),
            "taxable": round(after, 2)
        })
        subtotal += after
        savings += disc

    cgst = round(subtotal * 0.09, 2)
    sgst = round(subtotal * 0.09, 2)
    total = round(subtotal + cgst + sgst, 2)
    date_str = datetime.now().strftime("%d-%b-%Y")

    env = Environment(loader=FileSystemLoader("templates"))
    template = env.get_template("quotation_pdf.html")

    html = template.render(
        title="PROFORMA INVOICE",
        label="Proforma Invoice No:",
        invoice_no="Draft",
        items=items,
        subtotal=round(subtotal, 2),
        cgst=cgst,
        sgst=sgst,
        total=total,
        savings=round(savings, 2),
        date=date_str,
        buyer_name=customer_name,
        buyer_address=customer_address,
        buyer_gstin=customer_gstin,
        buyer_mobile=customer_mobile,
        buyer_email=customer_email
    )
    return HTMLResponse(content=html)


# Helper function to generate beautifully formatted single document Excel
def generate_single_doc_excel(
    title, doc_no_label, doc_no, date_val,
    buyer_name, buyer_address, buyer_gstin, buyer_mobile, buyer_email,
    items, is_quotation
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = title[:30]
    ws.views.sheetView[0].showGridLines = True

    # Colors and Fonts
    font_family = "Segoe UI"
    color_primary = "1E40AF"  # Mahindra blue
    color_text = "1F2937"
    color_success = "15803D"

    font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_bold = Font(name=font_family, size=10, bold=True, color=color_text)
    font_regular = Font(name=font_family, size=10, color=color_text)
    font_savings = Font(name=font_family, size=11, bold=True, color=color_success)

    fill_title = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
    fill_header = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    fill_light = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    fill_savings = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="D1D5DB")
    border_thin = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_double_bottom = Border(bottom=Side(style="double", color="374151"), top=thin_border_side)

    # Title Block
    ws.merge_cells("A1:G2")
    title_cell = ws["A1"]
    title_cell.value = f"MAHINDRA PRO SPARES - {title}"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Company Details & Meta Details
    ws["A4"] = "MAHINDRA PRO SPARES"
    ws["A4"].font = font_bold
    ws["A5"] = "SHOP NO.01, CIDCO BUILDING, NEAR DEVANSHI HOTEL,"
    ws["A5"].font = font_regular
    ws["A6"] = "TRUCK TERMINAL, KALAMBOLI, PANVEL, RAIGAD, MH 410218"
    ws["A6"].font = font_regular
    ws["A7"] = "GSTIN: 27BHIPM7720B1ZH | Contact: +91-8652369863"
    ws["A7"].font = font_regular

    ws["F4"] = f"{doc_no_label}:"
    ws["F4"].font = font_bold
    ws["G4"] = doc_no
    ws["G4"].font = font_bold
    ws["F5"] = "Date:"
    ws["F5"].font = font_bold
    ws["G5"] = date_val.strftime("%d-%b-%Y") if date_val else ""
    ws["G5"].font = font_regular

    # Customer Details Block
    ws["A9"] = "Bill To:"
    ws["A9"].font = font_bold
    ws["A10"] = buyer_name or "Walk-In Customer"
    ws["A10"].font = font_bold
    ws["A11"] = buyer_address or "—"
    ws["A11"].font = font_regular
    ws["A12"] = f"GSTIN: {buyer_gstin}" if buyer_gstin else ""
    ws["A12"].font = font_regular
    ws["A13"] = f"Mobile: {buyer_mobile} | Email: {buyer_email}" if (buyer_mobile or buyer_email) else ""
    ws["A13"].font = font_regular

    # Items Headers
    headers = ["Sl", "Part No / Description", "HSN", "Qty", "Rate (₹)", "Disc %", "Amount (₹)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=15, column=col_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if col_idx not in [2] else "left", vertical="center")
    ws.row_dimensions[15].height = 25

    # Item rows
    row_idx = 16
    subtotal = 0.0
    savings = 0.0
    for idx, item in enumerate(items, 1):
        ws.cell(row=row_idx, column=1, value=idx).alignment = Alignment(horizontal="center")
        
        part_desc = f"{item.part_no}\n{item.description}" if item.description else item.part_no
        cell_desc = ws.cell(row=row_idx, column=2, value=part_desc)
        cell_desc.alignment = Alignment(wrap_text=True, vertical="center")
        
        ws.cell(row=row_idx, column=3, value=item.hsn or "—").alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=4, value=int(item.qty)).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=5, value=float(item.rate)).number_format = "₹#,##0.00"
        
        disc = getattr(item, 'discount', 0.0) or 0.0
        ws.cell(row=row_idx, column=6, value=disc).alignment = Alignment(horizontal="center")
        
        amt = float(item.amount) if hasattr(item, 'amount') else (float(item.rate) * float(item.qty) * (1 - disc / 100))
        ws.cell(row=row_idx, column=7, value=amt).number_format = "₹#,##0.00"
        
        subtotal += amt
        savings += (float(item.rate) * float(item.qty) * (disc / 100))

        for c in range(1, 8):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = font_regular
            cell.border = border_thin
            if idx % 2 == 0:
                cell.fill = fill_light

        # Adjust height based on newline
        lines = part_desc.count('\n') + 1
        ws.row_dimensions[row_idx].height = 18 * lines
        row_idx += 1

    # Financial Summary
    cgst = round(subtotal * 0.09, 2)
    sgst = round(subtotal * 0.09, 2)
    grand_total = round(subtotal + cgst + sgst, 2)

    summary_rows = [
        ("Subtotal", subtotal),
        ("CGST (9%)", cgst),
        ("SGST (9%)", sgst),
        ("Grand Total", grand_total),
    ]

    for label, val in summary_rows:
        ws.cell(row=row_idx, column=6, value=label).font = font_bold if label == "Grand Total" else font_regular
        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right")
        
        val_cell = ws.cell(row=row_idx, column=7, value=val)
        val_cell.font = font_bold if label == "Grand Total" else font_regular
        val_cell.number_format = "₹#,##0.00"
        
        if label == "Grand Total":
            val_cell.border = border_double_bottom
            ws.cell(row=row_idx, column=6).border = Border(top=thin_border_side)
        row_idx += 1

    # Savings Star
    if savings > 0.01:
        row_idx += 1
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx+1, end_column=7)
        savings_cell = ws.cell(row=row_idx, column=1)
        savings_cell.value = f"★ You have saved {format_inr(savings)} from this order!"
        savings_cell.font = font_savings
        savings_cell.fill = fill_savings
        savings_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        thin_green_side = Side(border_style="dashDot", color="16A34A")
        green_border = Border(left=thin_green_side, right=thin_green_side, top=thin_green_side, bottom=thin_green_side)
        for r in range(row_idx, row_idx+2):
            for c in range(1, 8):
                ws.cell(row=r, column=c).border = green_border
        row_idx += 2

    # Set column widths
    column_widths = [6, 32, 10, 8, 14, 14, 16]
    for i, w in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{title.lower().replace(' ', '_')}_{doc_no}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def safe_format_date(d):
    if not d:
        return ""
    if isinstance(d, str):
        return d.split(" ")[0]
    try:
        return d.strftime("%d-%b-%Y")
    except Exception:
        return str(d)


def convert_html_to_pdf(html_content: str) -> bytes:
    from xhtml2pdf import pisa
    import io
    result = io.BytesIO()
    pisa_status = pisa.CreatePDF(io.BytesIO(html_content.encode("utf-8")), result)
    if not pisa_status.err:
        return result.getvalue()
    raise ValueError(f"xhtml2pdf error: {pisa_status.err}")


def safe_format_datetime(dt):
    if not dt:
        return ""
    if isinstance(dt, str):
        return dt.split(".")[0]
    try:
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(dt)


@app.get("/view_quotation_saved/{q_id}", response_class=HTMLResponse)
def view_quotation_saved(q_id: int):
    db = SessionLocal()
    quotation = db.execute(text("""
        SELECT id, quotation_no, customer_name, customer_address, customer_gstin, customer_mobile, customer_email, date
        FROM quotations WHERE id = :id
    """), {"id": q_id}).fetchone()
    
    if not quotation:
        db.close()
        return HTMLResponse("<h1>Proforma Invoice not found</h1>", status_code=404)
        
    items_rows = db.execute(text("""
        SELECT part_no, description, rate, qty, discount, amount, hsn
        FROM quotation_items WHERE quotation_id = :qid
    """), {"qid": q_id}).fetchall()
    db.close()

    items = []
    subtotal = 0
    savings = 0.0
    for r in items_rows:
        sub = r.rate * r.qty
        disc = sub * ((r.discount or 0) / 100)
        taxable = sub - disc
        items.append({
            "part_no": r.part_no,
            "description": r.description,
            "hsn": r.hsn,
            "rate": r.rate,
            "qty": r.qty,
            "discount": r.discount or 0,
            "taxable": round(taxable, 2)
        })
        subtotal += taxable
        savings += disc

    cgst = round(subtotal * 0.09, 2)
    sgst = round(subtotal * 0.09, 2)
    total = round(subtotal + cgst + sgst, 2)
    
    env = Environment(loader=FileSystemLoader("templates"))
    template = env.get_template("quotation_pdf.html")

    html = template.render(
        title="PROFORMA INVOICE",
        label="Proforma Invoice No:",
        invoice_no=quotation.quotation_no,
        items=items,
        subtotal=round(subtotal, 2),
        cgst=cgst,
        sgst=sgst,
        total=total,
        savings=round(savings, 2),
        date=safe_format_date(quotation.date),
        buyer_name=quotation.customer_name,
        buyer_address=quotation.customer_address,
        buyer_gstin=quotation.customer_gstin,
        buyer_mobile=quotation.customer_mobile,
        buyer_email=quotation.customer_email
    )
    return HTMLResponse(content=html)


@app.get("/download_quotation_pdf_file/{q_id}")
def download_quotation_pdf_file(q_id: int):
    db = SessionLocal()
    quotation = db.execute(text("""
        SELECT id, quotation_no, customer_name, customer_address, customer_gstin, customer_mobile, customer_email, date
        FROM quotations WHERE id = :id
    """), {"id": q_id}).fetchone()
    
    if not quotation:
        db.close()
        return HTMLResponse("<h1>Proforma Invoice not found</h1>", status_code=404)
        
    items_rows = db.execute(text("""
        SELECT part_no, description, rate, qty, discount, amount, hsn
        FROM quotation_items WHERE quotation_id = :qid
    """), {"qid": q_id}).fetchall()
    db.close()

    items = []
    subtotal = 0
    savings = 0.0
    for r in items_rows:
        sub = r.rate * r.qty
        disc = sub * ((r.discount or 0) / 100)
        taxable = sub - disc
        items.append({
            "part_no": r.part_no,
            "description": r.description,
            "hsn": r.hsn,
            "rate": r.rate,
            "qty": r.qty,
            "discount": r.discount or 0,
            "taxable": round(taxable, 2)
        })
        subtotal += taxable
        savings += disc

    cgst = round(subtotal * 0.09, 2)
    sgst = round(subtotal * 0.09, 2)
    total = round(subtotal + cgst + sgst, 2)
    
    env = Environment(loader=FileSystemLoader("templates"))
    template = env.get_template("quotation_pdf.html")

    html = template.render(
        title="PROFORMA INVOICE",
        label="Proforma Invoice No:",
        invoice_no=quotation.quotation_no,
        items=items,
        subtotal=round(subtotal, 2),
        cgst=cgst,
        sgst=sgst,
        total=total,
        savings=round(savings, 2),
        date=safe_format_date(quotation.date),
        buyer_name=quotation.customer_name,
        buyer_address=quotation.customer_address,
        buyer_gstin=quotation.customer_gstin,
        buyer_mobile=quotation.customer_mobile,
        buyer_email=quotation.customer_email
    )
    
    try:
        pdf_bytes = convert_html_to_pdf(html)
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=proforma_{quotation.quotation_no}.pdf"}
        )
    except Exception as e:
        fallback_html = html + "\n<script>window.onload = function() { window.print(); }</script>"
        return HTMLResponse(content=fallback_html)


def send_smtp_email(to_email: str, subject: str, html_body: str, attachment_bytes: bytes = None, attachment_name: str = ""):
    smtp_host = os.environ.get("SMTP_HOST", "").strip()
    smtp_port = os.environ.get("SMTP_PORT", "").strip()
    smtp_user = os.environ.get("SMTP_USER", "").strip()
    smtp_password = os.environ.get("SMTP_PASSWORD", "").strip()
    smtp_from = os.environ.get("SMTP_FROM", smtp_user).strip()
    
    if not smtp_host or not smtp_user or not smtp_password:
        raise ValueError("SMTP_NOT_CONFIGURED")
        
    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"] = smtp_from
    msg["To"] = to_email
    
    body_part = MIMEText(html_body, "html")
    msg.attach(body_part)
    
    if attachment_bytes:
        from email.mime.application import MIMEApplication
        att = MIMEApplication(attachment_bytes, _subtype="pdf")
        att.add_header('Content-Disposition', 'attachment', filename=attachment_name)
        msg.attach(att)
        
    port = int(smtp_port) if smtp_port.isdigit() else 587
    if port == 465:
        server = smtplib.SMTP_SSL(smtp_host, port)
    else:
        server = smtplib.SMTP(smtp_host, port)
        server.starttls()
        
    server.login(smtp_user, smtp_password)
    server.sendmail(smtp_from, to_email, msg.as_string())
    server.quit()


@app.post("/send_email_quotation/{q_id}")
async def send_email_quotation(q_id: int, request: Request, email: str = Query(None)):
    db = SessionLocal()
    quotation = db.execute(text("""
        SELECT id, quotation_no, customer_name, customer_address, customer_gstin, customer_mobile, customer_email, date
        FROM quotations WHERE id = :id
    """), {"id": q_id}).fetchone()
    
    if not quotation:
        db.close()
        return {"ok": False, "error": "Proforma Invoice not found"}
        
    to_email = email.strip() if email else (quotation.customer_email or "").strip()
    if not to_email:
        db.close()
        return {"ok": False, "error": "CUSTOMER_EMAIL_MISSING"}
        
    items_rows = db.execute(text("""
        SELECT part_no, description, rate, qty, discount, amount, hsn
        FROM quotation_items WHERE quotation_id = :qid
    """), {"qid": q_id}).fetchall()
    db.close()

    items = []
    subtotal = 0
    savings = 0.0
    for r in items_rows:
        sub = r.rate * r.qty
        disc = sub * ((r.discount or 0) / 100)
        taxable = sub - disc
        items.append({
            "part_no": r.part_no,
            "description": r.description,
            "hsn": r.hsn,
            "rate": r.rate,
            "qty": r.qty,
            "discount": r.discount or 0,
            "taxable": round(taxable, 2)
        })
        subtotal += taxable
        savings += disc

    cgst = round(subtotal * 0.09, 2)
    sgst = round(subtotal * 0.09, 2)
    total = round(subtotal + cgst + sgst, 2)
    
    env = Environment(loader=FileSystemLoader("templates"))
    template = env.get_template("quotation_pdf.html")

    html = template.render(
        title="PROFORMA INVOICE",
        label="Proforma Invoice No:",
        invoice_no=quotation.quotation_no,
        items=items,
        subtotal=round(subtotal, 2),
        cgst=cgst,
        sgst=sgst,
        total=total,
        savings=round(savings, 2),
        date=safe_format_date(quotation.date),
        buyer_name=quotation.customer_name,
        buyer_address=quotation.customer_address,
        buyer_gstin=quotation.customer_gstin,
        buyer_mobile=quotation.customer_mobile,
        buyer_email=quotation.customer_email
    )
    
    # Try generating PDF to attach to email
    pdf_bytes = None
    try:
        pdf_bytes = convert_html_to_pdf(html)
    except Exception:
        pass
        
    try:
        subject = f"Proforma Invoice {quotation.quotation_no} - Mahindra Pro Spares"
        email_body = f"""
        <p>Dear Customer,</p>
        <p>Please find attached your Proforma Invoice <b>{quotation.quotation_no}</b> from Mahindra Pro Spares.</p>
        <br>
        <p>Grand Total: ₹ {total:,.2f}</p>
        <p>Date: {safe_format_date(quotation.date)}</p>
        <br>
        <p>Thank you for your business!</p>
        """
        send_smtp_email(
            to_email=to_email,
            subject=subject,
            html_body=email_body,
            attachment_bytes=pdf_bytes,
            attachment_name=f"proforma_{quotation.quotation_no}.pdf"
        )
        return {"ok": True}
    except ValueError as ve:
        if str(ve) == "SMTP_NOT_CONFIGURED":
            return {"ok": False, "error": "SMTP_NOT_CONFIGURED"}
        return {"ok": False, "error": str(ve)}
    except Exception as e:
        return {"ok": False, "error": f"Failed to send email: {str(e)}"}


@app.get("/view_invoice_saved/{i_id}", response_class=HTMLResponse)
def view_invoice_saved(i_id: int):
    db = SessionLocal()
    invoice = db.execute(text("""
        SELECT id, invoice_no, customer_name, customer_address, customer_gstin, customer_mobile, customer_email, date
        FROM invoices WHERE id = :id
    """), {"id": i_id}).fetchone()
    
    if not invoice:
        db.close()
        return HTMLResponse("<h1>Tax Invoice not found</h1>", status_code=404)
        
    items_rows = db.execute(text("""
        SELECT part_no, description, rate, quantity, amount, hsn
        FROM invoice_items WHERE invoice_id = :iid
    """), {"iid": i_id}).fetchall()
    db.close()

    items = []
    subtotal = 0
    savings = 0.0
    for r in items_rows:
        sub = r.rate * r.quantity
        taxable = r.amount
        disc_amt = max(0.0, sub - taxable)
        disc_pct = round((disc_amt / sub) * 100, 2) if sub > 0 else 0.0
        items.append({
            "part_no": r.part_no,
            "description": r.description,
            "hsn": r.hsn,
            "rate": r.rate,
            "qty": r.quantity,
            "discount": disc_pct,
            "taxable": round(taxable, 2)
        })
        subtotal += taxable
        savings += disc_amt

    cgst = round(subtotal * 0.09, 2)
    sgst = round(subtotal * 0.09, 2)
    total = round(subtotal + cgst + sgst, 2)
    
    env = Environment(loader=FileSystemLoader("templates"))
    template = env.get_template("quotation_pdf.html")

    html = template.render(
        title="TAX INVOICE",
        label="Invoice No:",
        invoice_no=invoice.invoice_no,
        items=items,
        subtotal=round(subtotal, 2),
        cgst=cgst,
        sgst=sgst,
        total=total,
        savings=round(savings, 2),
        date=safe_format_date(invoice.date),
        buyer_name=invoice.customer_name,
        buyer_address=invoice.customer_address,
        buyer_gstin=invoice.customer_gstin,
        buyer_mobile=invoice.customer_mobile,
        buyer_email=invoice.customer_email
    )
    return HTMLResponse(content=html)


@app.get("/export_quotation_excel/{q_id}")
def export_quotation_excel(q_id: int):
    db = SessionLocal()
    quotation = db.execute(text("""
        SELECT id, quotation_no, customer_name, customer_address, customer_gstin, customer_mobile, customer_email, date
        FROM quotations WHERE id = :id
    """), {"id": q_id}).fetchone()
    
    if not quotation:
        db.close()
        return HTMLResponse("<h1>Proforma Invoice not found</h1>", status_code=404)
        
    items_rows = db.execute(text("""
        SELECT part_no, description, rate, qty, discount, amount, hsn
        FROM quotation_items WHERE quotation_id = :qid
    """), {"qid": q_id}).fetchall()
    db.close()

    return generate_single_doc_excel(
        title="PROFORMA INVOICE",
        doc_no_label="Proforma Invoice No",
        doc_no=quotation.quotation_no,
        date_val=quotation.date,
        buyer_name=quotation.customer_name,
        buyer_address=quotation.customer_address,
        buyer_gstin=quotation.customer_gstin,
        buyer_mobile=quotation.customer_mobile,
        buyer_email=quotation.customer_email,
        items=items_rows,
        is_quotation=True
    )


@app.get("/export_invoice_excel/{i_id}")
def export_invoice_excel(i_id: int):
    db = SessionLocal()
    invoice = db.execute(text("""
        SELECT id, invoice_no, customer_name, customer_address, customer_gstin, customer_mobile, customer_email, date
        FROM invoices WHERE id = :id
    """), {"id": i_id}).fetchone()
    
    if not invoice:
        db.close()
        return HTMLResponse("<h1>Invoice not found</h1>", status_code=404)
        
    items_rows = db.execute(text("""
        SELECT part_no, description, rate, quantity as qty, amount, hsn
        FROM invoice_items WHERE invoice_id = :iid
    """), {"iid": i_id}).fetchall()
    db.close()

    # Adapt format
    adapted_items = []
    for r in items_rows:
        sub = r.rate * r.qty
        disc_amt = max(0.0, sub - r.amount)
        disc_pct = (disc_amt / sub) * 100 if sub > 0 else 0.0
        adapted_items.append(type('Item', (object,), {
            "part_no": r.part_no,
            "description": r.description,
            "hsn": r.hsn,
            "rate": r.rate,
            "qty": r.qty,
            "discount": disc_pct,
            "amount": r.amount
        }))

    return generate_single_doc_excel(
        title="TAX INVOICE",
        doc_no_label="Invoice No",
        doc_no=invoice.invoice_no,
        date_val=invoice.date,
        buyer_name=invoice.customer_name,
        buyer_address=invoice.customer_address,
        buyer_gstin=invoice.customer_gstin,
        buyer_mobile=invoice.customer_mobile,
        buyer_email=invoice.customer_email,
        items=adapted_items,
        is_quotation=False
    )


@app.get("/export_all_quotations_excel")
def export_all_quotations_excel(
    start_date: str,
    end_date: str,
    customer: str = Query("")
):
    db = SessionLocal()
    sql_parts = [
        "FROM quotation_items qi",
        "JOIN quotations q ON qi.quotation_id = q.id",
        "WHERE CAST(q.date AS DATE) BETWEEN :start AND :end"
    ]
    params = {
        "start": start_date,
        "end": end_date
    }
    if customer:
        sql_parts.append("AND LOWER(q.customer_name) LIKE LOWER(:customer)")
        params["customer"] = f"%{customer}%"

    query = text(f"""
        SELECT 
            q.quotation_no,
            q.customer_name,
            qi.part_no,
            qi.description,
            qi.qty,
            qi.rate,
            qi.discount,
            qi.amount as taxable_amount,
            q.date
        {chr(10).join(sql_parts)}
        ORDER BY q.date DESC
    """)

    result = db.execute(query, params).fetchall()
    db.close()

    rows = []
    for r in result:
        taxable = float(r.taxable_amount or 0.0)
        gst = taxable * 0.18
        cgst = gst / 2
        sgst = gst / 2
        grand = taxable + gst
        qty = float(r.qty or 0.0)
        rate = float(r.rate or 0.0)
        disc = float(r.discount or 0.0)
        saved = (qty * rate) - taxable

        rows.append({
            "Proforma Invoice No": r.quotation_no,
            "Customer": r.customer_name,
            "Part No": r.part_no,
            "Description": r.description,
            "Qty": qty,
            "Selling Rate": rate,
            "Discount %": disc,
            "Taxable Amount": round(taxable, 2),
            "CGST": round(cgst, 2),
            "SGST": round(sgst, 2),
            "Grand Total": round(grand, 2),
            "Saved Amount": round(saved, 2),
            "Date": safe_format_datetime(r.date)
        })

    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        "Proforma Invoice No", "Customer", "Part No", "Description", "Qty", "Selling Rate", "Discount %",
        "Taxable Amount", "CGST", "SGST", "Grand Total", "Saved Amount", "Date"
    ])

    total_orders = len(df["Proforma Invoice No"].unique()) if not df.empty else 0
    total_sales = df["Taxable Amount"].sum() if not df.empty else 0.0
    total_cgst = df["CGST"].sum() if not df.empty else 0.0
    total_sgst = df["SGST"].sum() if not df.empty else 0.0
    grand_total = df["Grand Total"].sum() if not df.empty else 0.0
    total_saved = df["Saved Amount"].sum() if not df.empty else 0.0

    summary_df = pd.DataFrame({
        "Metric": [
            "Total Proforma Orders",
            "Total Sales (Taxable)",
            "Total CGST",
            "Total SGST",
            "Grand Total (incl. GST)",
            "Total Saved (due to Discounts)"
        ],
        "Value": [
            total_orders,
            round(total_sales, 2),
            round(total_cgst, 2),
            round(total_sgst, 2),
            round(grand_total, 2),
            round(total_saved, 2)
        ]
    })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Proforma Data", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=proforma_sales.xlsx"}
    )

@app.get("/out_of_stock_products")
def get_out_of_stock_products(request: Request):
    active_store = get_active_store(request)
    db = SessionLocal()
    products = db.query(Product).filter(Product.store == active_store, Product.quantity == 0).all()
    db.close()
    return [{
        "part_no": p.part_no,
        "description": p.description,
        "hsn": p.hsn,
        "gst": p.gst or 18.0
    } for p in products]


@app.post("/download_order_book_pdf")
async def download_order_book_pdf(request: Request):
    data = await request.json()
    rows = data.get("rows", [])
    
    # Exclude rates and discounts as requested. Just show sl, part_no, description, hsn, gst, qty
    items = []
    for r in rows:
        items.append({
            "part_no": r.get("part_no", ""),
            "description": r.get("description", ""),
            "hsn": r.get("hsn", ""),
            "gst": r.get("gst", 18.0),
            "qty": r.get("qty", 1)
        })

    date_str = datetime.now().strftime("%d-%b-%Y")
    
    env = Environment(loader=FileSystemLoader("templates"))
    template = env.get_template("order_book_pdf.html")

    html = template.render(
        items=items,
        date=date_str
    )
    return HTMLResponse(content=html)


@app.post("/download_order_book_excel")
async def download_order_book_excel(request: Request):
    data = await request.json()
    rows = data.get("rows", [])
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Order Book"
    ws.views.sheetView[0].showGridLines = True

    # Colors and Fonts
    font_family = "Segoe UI"
    color_primary = "1E40AF"  # Mahindra blue
    color_text = "1F2937"

    font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_bold = Font(name=font_family, size=10, bold=True, color=color_text)
    font_regular = Font(name=font_family, size=10, color=color_text)

    fill_title = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
    fill_header = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    fill_light = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="D1D5DB")
    border_thin = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Title Block
    ws.merge_cells("A1:D2")
    title_cell = ws["A1"]
    title_cell.value = "MAHINDRA PRO SPARES - ORDER BOOK"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Company Details & Meta Details
    ws["A4"] = "MAHINDRA PRO SPARES"
    ws["A4"].font = font_bold
    ws["A5"] = "SHOP NO.01, CIDCO BUILDING, NEAR DEVANSHI HOTEL,"
    ws["A5"].font = font_regular
    ws["A6"] = "TRUCK TERMINAL, KALAMBOLI, PANVEL, RAIGAD, MH 410218"
    ws["A6"].font = font_regular

    ws["C4"] = "Document Type:"
    ws["C4"].font = font_bold
    ws["D4"] = "Order Book / Purchase Order"
    ws["D4"].font = font_regular
    ws["C5"] = "Date:"
    ws["C5"].font = font_bold
    ws["D5"] = datetime.now().strftime("%d-%b-%Y")
    ws["D5"].font = font_regular

    # Items Headers
    headers = ["Sl", "Part No / Description", "HSN", "Order Qty"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left", vertical="center")
    ws.row_dimensions[8].height = 25

    # Item rows
    row_idx = 9
    for idx, item in enumerate(rows, 1):
        ws.cell(row=row_idx, column=1, value=idx).alignment = Alignment(horizontal="center")
        
        part_no = item.get("part_no", "")
        desc = item.get("description", "")
        part_desc = f"{part_no}\n{desc}" if desc else part_no
        cell_desc = ws.cell(row=row_idx, column=2, value=part_desc)
        cell_desc.alignment = Alignment(wrap_text=True, vertical="center")
        
        ws.cell(row=row_idx, column=3, value=item.get("hsn", "") or "—").alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=4, value=int(item.get("qty", 1))).alignment = Alignment(horizontal="center")
        
        for c in range(1, 5):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = font_regular
            cell.border = border_thin
            if idx % 2 == 0:
                cell.fill = fill_light

        # Adjust height based on newline
        lines = part_desc.count('\n') + 1
        ws.row_dimensions[row_idx].height = 18 * lines
        row_idx += 1

    # Footer / Notes
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Notes:").font = font_bold
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="This is an order book summary generated for procuring out-of-stock inventory.").font = font_regular

    # Prepared By
    row_idx += 2
    ws.cell(row=row_idx, column=3, value="Prepared By:").font = font_bold
    ws.cell(row=row_idx, column=4, value="__________________").font = font_regular

    # Set column widths
    column_widths = [6, 45, 12, 14]
    for i, w in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)

    filename = f"order_book_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ----------------------------------------------------------------



def generate_invoice_no(db):
    import calendar
    today = datetime.now()

    year_month = today.strftime("%Y-%m")   # 2026-06

    # Python-side date range — works with both SQLite and PostgreSQL
    first_day = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_day_num = calendar.monthrange(today.year, today.month)[1]
    last_day = today.replace(day=last_day_num, hour=23, minute=59, second=59, microsecond=999999)

    count = db.execute(text("""
        SELECT COUNT(*) 
        FROM invoices 
        WHERE date >= :from_date AND date <= :to_date
    """), {"from_date": first_day, "to_date": last_day}).scalar()

    serial = str(count + 1).zfill(3)   # 001, 002, 003
    return f"{year_month}-{serial}"
@app.post("/download_pdf")
async def download_pdf(request: Request):

    payload = await request.json()
    # Support both flat list (old) and {rows:[], customer_name:""} (new)
    if isinstance(payload, list):
        data = payload
        customer_name = "Walk-In Customer"
        customer_address = ""
        customer_gstin = ""
        customer_mobile = ""
        customer_email = ""
    else:
        data = payload.get("rows", payload) if isinstance(payload.get("rows"), list) else payload
        customer_name = payload.get("customer_name", "").strip() or "Walk-In Customer"
        customer_address = payload.get("customer_address", "").strip() or ""
        customer_gstin = payload.get("customer_gstin", "").strip() or ""
        customer_mobile = payload.get("customer_mobile", "").strip() or ""
        customer_email = payload.get("customer_email", "").strip() or ""
        if isinstance(data, dict):  # still not right, fall back
            data = [payload] if "part_no" in payload else []
    db = SessionLocal()
    from collections import defaultdict

    try:
        # ===============================
        # ✅ STEP 1: VALIDATE STOCK
        # ===============================
        qty_map = defaultdict(float)

        for row in data:
            qty_map[row["part_no"]] += float(row["qty"])

        for part_no, total_qty in qty_map.items():
            product = db.query(Product).filter(Product.part_no == part_no).first()

            if not product or product.quantity < total_qty:
                raise ValueError(f"Not enough stock for {part_no}")

        # ===============================
        # ✅ Save Vendor Details
        # ===============================
        if customer_name and customer_name != "Walk-In Customer":
            c_name = customer_name
            existing_vendor = db.query(Vendor).filter(Vendor.name == c_name).first()
            if existing_vendor:
                if customer_address: existing_vendor.address = customer_address
                if customer_mobile: existing_vendor.mobile_num = customer_mobile
                if customer_gstin: existing_vendor.gstin = customer_gstin
                if customer_email: existing_vendor.email_id = customer_email
            else:
                new_vendor = Vendor(
                    name=c_name,
                    address=customer_address,
                    mobile_num=customer_mobile,
                    gstin=customer_gstin,
                    email_id=customer_email
                )
                db.add(new_vendor)
            db.commit()

        # ===============================
        # ✅ STEP 2: PROCESS + REDUCE STOCK
        # ===============================
        items = []
        subtotal = 0
        active_store = get_active_store(request)

        for row in data:
            product = db.query(Product).filter(Product.part_no == row["part_no"], Product.store == active_store).first()
            db_item = db.execute(
                text("SELECT description, mrp, hsn FROM order_items WHERE part_no=:p AND store=:store"),
                {"p": row["part_no"], "store": active_store}
            ).fetchone()

            desc = row.get("description", "") or (product.description if product else "") or (db_item.description if db_item else "")
            hsn = row.get("hsn", "") or (product.hsn if product else "") or (str(db_item.hsn) if db_item else "")

            # Rate Priority: 1. UI entered rate, 2. Stock Data rate, 3. Price Master MRP
            rate = float(row.get("rate", 0))
            if rate == 0 and product and product.rate and float(product.rate) > 0:
                rate = float(product.rate)
            if rate == 0 and db_item and db_item.mrp and float(db_item.mrp) > 0:
                rate = float(db_item.mrp)

            qty  = float(row.get("qty", 1))
            disc = float(row.get("discount", 0))

            purchase_rate = 0.0
            if product:
                purchase_rate = float(product.rate or 0.0)
                product.quantity -= qty

                if product.quantity <= 0:
                    product.quantity = 0
                    product.amount = 0
                else:
                    product.amount = (product.quantity * (product.rate or 0)) - (
                        (product.quantity * (product.rate or 0)) * ((product.discount or 0) / 100)
                    )

            # 🔥 CALCULATIONS
            base = qty * rate
            discount_amt = base * (disc / 100)
            taxable = base - discount_amt

            subtotal += taxable

            items.append({
                "part_no": row["part_no"],
                "description": desc,
                "hsn": hsn,
                "qty": qty,
                "rate": rate,
                "discount": disc,
                "taxable": round(taxable, 2),
                "purchase_rate": purchase_rate
            })

        # ===============================
        # ✅ STEP 3: GST
        # ===============================
        cgst = round(subtotal * 0.09, 2)
        sgst = round(subtotal * 0.09, 2)
        total = round(subtotal + cgst + sgst, 2)

        # ===============================
        # ✅ STEP 4: SAVE INVOICE
        # ===============================
        invoice_no = generate_invoice_no(db)

        # Use PostgreSQL-compatible insert with RETURNING, fallback to lastrowid
        from database import engine as _engine
        is_postgres = "postgresql" in str(_engine.url)

        if is_postgres:
            result = db.execute(text("""
                INSERT INTO invoices (
                    invoice_no, customer_name, customer_address, customer_gstin, customer_mobile, customer_email, date,
                    total_amount, gst_amount, grand_total, store
                ) VALUES (
                    :inv, :cust, :address, :gstin, :mobile, :email, :date,
                    :total, :gst, :grand, :store
                ) RETURNING id
            """), {
                "inv": invoice_no, "cust": customer_name,
                "address": customer_address, "gstin": customer_gstin,
                "mobile": customer_mobile, "email": customer_email,
                "date": datetime.utcnow(),
                "total": round(subtotal, 2),
                "gst": cgst + sgst, "grand": total,
                "store": active_store
            })
            invoice_id = result.fetchone()[0]
        else:
            result = db.execute(text("""
                INSERT INTO invoices (
                    invoice_no, customer_name, customer_address, customer_gstin, customer_mobile, customer_email, date,
                    total_amount, gst_amount, grand_total, store
                ) VALUES (
                    :inv, :cust, :address, :gstin, :mobile, :email, :date,
                    :total, :gst, :grand, :store
                )
            """), {
                "inv": invoice_no, "cust": customer_name,
                "address": customer_address, "gstin": customer_gstin,
                "mobile": customer_mobile, "email": customer_email,
                "date": datetime.utcnow(),
                "total": round(subtotal, 2),
                "gst": cgst + sgst, "grand": total,
                "store": active_store
            })
            invoice_id = result.lastrowid

        for it in items:
            db.execute(text("""
                INSERT INTO invoice_items 
                (invoice_id, part_no, description, quantity, rate, amount, hsn, purchase_rate)
                VALUES (:iid, :p, :d, :q, :r, :amt, :hsn, :prate)
            """), {
                "iid": invoice_id,
                "p": it["part_no"],
                "d": it["description"],
                "q": it["qty"],
                "r": it["rate"],
                "amt": it["taxable"],
                "hsn": it["hsn"],
                "prate": it["purchase_rate"]
            })

        db.commit()

    except Exception as e:
        db.rollback()
        db.close()
        raise e

    db.close()

    # ===============================
    # ✅ STEP 5: GENERATE PDF
    # ===============================
    env = Environment(loader=FileSystemLoader("templates"))
    template = env.get_template("quotation_pdf.html")

    savings = sum((it["rate"] * it["qty"]) - it["taxable"] for it in items)

    html = template.render(
        title="TAX INVOICE",
        label="Invoice No:",
        invoice_no=invoice_no,
        items=items,
        subtotal=round(subtotal, 2),
        cgst=cgst,
        sgst=sgst,
        total=total,
        savings=round(savings, 2),
        date=datetime.now().strftime("%d-%b-%Y"),
        buyer_name=customer_name,
        buyer_address=customer_address,
        buyer_gstin=customer_gstin,
        buyer_mobile=customer_mobile,
        buyer_email=customer_email
    )

    from fastapi.responses import HTMLResponse
    return HTMLResponse(content=html)

from fastapi import Query
from datetime import datetime

@app.get("/sales_summary")
def sales_summary(
    request: Request,
    start_date: str = Query(...),
    end_date: str = Query(...),
    customer: str = Query(""),
    vendor: str = Query("")
):
    active_store = get_active_store(request)
    db = SessionLocal()

    sql_parts = [
        "FROM invoice_items ii",
        "JOIN invoices i ON ii.invoice_id = i.id",
        "LEFT JOIN products p ON (ii.part_no = p.part_no AND p.store = :store)",
        "WHERE i.store = :store AND CAST(i.date AS DATE) BETWEEN :start AND :end"
    ]
    
    params = {
        "start": start_date,
        "end": end_date,
        "store": active_store
    }
    
    if customer:
        sql_parts.append("AND LOWER(i.customer_name) LIKE LOWER(:customer)")
        params["customer"] = f"%{customer}%"
        
    if vendor:
        sql_parts.append("AND p.vendor_name = :vendor")
        params["vendor"] = vendor

    query_str = f"""
        SELECT 
            COUNT(DISTINCT i.id) as total_orders,
            SUM(ii.amount) as total_sales,
            SUM(COALESCE(ii.purchase_rate, 0.0) * ii.quantity) as total_cost
        {chr(10).join(sql_parts)}
    """
    
    result = db.execute(text(query_str), params).fetchone()
    
    total_sales = float(result.total_sales or 0.0)
    total_cost = float(result.total_cost or 0.0)
    total_gst = total_sales * 0.18
    grand_total = total_sales + total_gst
    total_profit = total_sales - total_cost

    # Fetch detailed sales list
    items_query = text(f"""
        SELECT 
            i.id as invoice_id,
            i.invoice_no,
            i.customer_name,
            ii.part_no,
            ii.description,
            ii.quantity,
            ii.rate,
            ii.amount as taxable_amount,
            COALESCE(ii.purchase_rate, 0.0) as purchase_rate,
            i.date
        {chr(10).join(sql_parts)}
        ORDER BY i.date DESC
    """)
    items_result = db.execute(items_query, params).fetchall()
    
    items_list = []
    for r in items_result:
        taxable = float(r.taxable_amount or 0.0)
        gst = taxable * 0.18
        cgst = gst / 2
        sgst = gst / 2
        grand = taxable + gst
        qty = float(r.quantity or 0.0)
        prate = float(r.purchase_rate or 0.0)
        pcost = prate * qty
        profit = taxable - pcost
        items_list.append({
            "invoice_id": r.invoice_id,
            "date": safe_format_datetime(r.date),
            "invoice_no": r.invoice_no,
            "customer_name": r.customer_name,
            "part_no": r.part_no,
            "description": r.description,
            "qty": qty,
            "rate": round(r.rate, 2) if r.rate else 0.0,
            "taxable": round(taxable, 2),
            "cgst": round(cgst, 2),
            "sgst": round(sgst, 2),
            "grand_total": round(grand, 2),
            "purchase_rate": round(prate, 2),
            "purchase_cost": round(pcost, 2),
            "profit": round(profit, 2)
        })

    db.close()
    
    return {
        "orders": result.total_orders or 0,
        "sales": round(total_sales, 2),
        "gst": round(total_gst, 2),
        "grand_total": round(grand_total, 2),
        "cost": round(total_cost, 2),
        "profit": round(total_profit, 2),
        "items": items_list
    }


@app.get("/purchase_summary")
def purchase_summary(
    request: Request,
    start_date: str = Query(...),
    end_date: str = Query(...),
    vendor: str = Query("")
):
    active_store = get_active_store(request)
    db = SessionLocal()
    
    summary_query = """
        SELECT 
            COUNT(*) as total_purchases,
            SUM(amount) as total_amount
        FROM purchases
        WHERE store = :store AND CAST(date AS DATE) BETWEEN :start AND :end
    """
    
    items_query = """
        SELECT 
            id, vendor_name, part_no, description, hsn, quantity, rate, discount, amount, date
        FROM purchases
        WHERE store = :store AND CAST(date AS DATE) BETWEEN :start AND :end
    """
    
    params = {
        "start": start_date,
        "end": end_date,
        "store": active_store
    }
    
    if vendor.strip():
        summary_query += " AND UPPER(vendor_name) = :vendor"
        items_query += " AND UPPER(vendor_name) = :vendor"
        params["vendor"] = vendor.strip().upper()
        
    items_query += " ORDER BY date DESC"
    
    summary_result = db.execute(text(summary_query), params).fetchone()
    items_result = db.execute(text(items_query), params).fetchall()
    db.close()
    
    return {
        "count": summary_result.total_purchases or 0,
        "total_amount": float(summary_result.total_amount or 0.0),
        "items": [{
            "id": r.id,
            "vendor_name": r.vendor_name,
            "part_no": r.part_no,
            "description": r.description,
            "hsn": r.hsn,
            "quantity": r.quantity,
            "rate": float(r.rate or 0.0),
            "discount": float(r.discount or 0.0),
            "amount": float(r.amount or 0.0),
            "date": safe_format_datetime(r.date)
        } for r in items_result]
    }


@app.get("/export_purchase_excel")
def export_purchase_excel(
    request: Request,
    start_date: str = Query(...),
    end_date: str = Query(...),
    vendor: str = Query("")
):
    active_store = get_active_store(request)
    db = SessionLocal()
    
    items_query = """
        SELECT 
            date, vendor_name, part_no, description, hsn, quantity, rate, discount, amount
        FROM purchases
        WHERE store = :store AND CAST(date AS DATE) BETWEEN :start AND :end
    """
    
    params = {
        "start": start_date,
        "end": end_date,
        "store": active_store
    }
    
    if vendor.strip():
        items_query += " AND UPPER(vendor_name) = :vendor"
        params["vendor"] = vendor.strip().upper()
        
    items_query += " ORDER BY date DESC"
    
    items_result = db.execute(text(items_query), params).fetchall()
    db.close()
    
    df = pd.DataFrame(items_result, columns=[
        "Date", "Vendor Name", "Part No", "Description", "HSN", "Quantity", "Rate", "Discount %", "Total Amount"
    ])
    
    if not df.empty and "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%Y-%m-%d %H:%M:%S")
        
    total_orders = len(df)
    total_amount = df["Total Amount"].sum() if not df.empty else 0.0
    
    summary_df = pd.DataFrame({
        "Metric": [
            "Total Purchase Transactions",
            "Total Purchase Value"
        ],
        "Value": [
            total_orders,
            total_amount
        ]
    })
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Purchase Data", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        
    output.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="purchase_report_{start_date}_to_{end_date}.xlsx"'
    }
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


@app.get("/import_pdf_data")
def import_pdf_data():
    db = SessionLocal()

    # 🔥 call your existing function
    products = extract_products("products.pdf")  

    for p in products:
        db.add(Product(
            part_no=p.get("part_no"),
            description=p.get("description"),
            hsn=p.get("hsn"),
            gst=float(p.get("gst", 18)),
            quantity=int(p.get("qty", 0)),
            rate=float(p.get("rate", 0)),
            discount=float(p.get("discount", 0)),
            amount=float(p.get("amount", 0))
        ))

    db.commit()
    db.close()

    return {"message": "PDF imported successfully"}
@app.get("/sales")
def get_sales(start: str, end: str):
    db = SessionLocal()

    rows = db.execute(text("""
        SELECT * FROM invoices
        WHERE CAST(date AS DATE) BETWEEN :s AND :e
        ORDER BY date DESC
    """), {
        "s": start,
        "e": end
    }).fetchall()

    total_sales = sum([r.total for r in rows])

    db.close()

    return {
        "total_sales": total_sales,
        "count": len(rows),
        "data": [dict(r._mapping) for r in rows]
    }

@app.post("/view_quotation", response_class=HTMLResponse)
async def view_quotation(request: Request):

    data = await request.json()

    items = []
    total = 0

    for item in data:
        subtotal = item["rate"] * item["qty"]
        discount = subtotal * (item["discount"] / 100)
        after = subtotal - discount
        gst = after * 0.18
        final = after + gst

        total += final

    
        db = SessionLocal()

     
        price = db.query(OrderItems).filter(
            OrderItems.part_no == item["part_no"]
        ).first()

        description = price.description if price else ""
        hsn = price.hsn if price else ""
        
        items.append({
            "part_no": item["part_no"],
            "description": description,
            "hsn": hsn,
            "qty": item["qty"],
            "rate": item["rate"],
            "discount": item["discount"],
            "final": round(final, 2)
        })

    return render(
        "quotation.html",
        items=items,
        total=round(total, 2),
        date=datetime.now().strftime("%d-%b-%Y"),
        buyer_name="Thakur Infraprojects Private Limited",
        buyer_address="""Plot No. 265/01, Om Sadanika, Panvel
Uran Road Uran Naka, Panvel, Raigad
GSTIN: 27AACCT6451F1ZC"""
    )
@app.get("/build_quotation")
def build_quotation():
    db = SessionLocal()

    rows = db.execute(text("SELECT * FROM order_items"))
    result = []

    for r in rows:
        stock = db.query(Product).filter(
            Product.part_no == r.part_no
        ).first()

        result.append({
            "part_no": r.part_no,
            "description": r.description,
            "rate": float(r.mrp),   # ✅ FIX
            "stock": stock.quantity if stock else 0,
            "qty": stock.quantity if stock else 0,
            "discount": 0
        })

    db.close()
    return result

@app.post("/upload_order")
def upload_order(file: UploadFile = File(...)):

    if file.filename.endswith(".csv"):
        df = pd.read_csv(file.file)
    else:
        df = pd.read_excel(file.file, engine="openpyxl")

    df.columns = [col.strip().upper() for col in df.columns]

    db = SessionLocal()

    db.execute(text("DELETE FROM order_items"))

    df = df.fillna("")

    for _, row in df.iterrows():

        part_no = str(row.get("PART NO", "")).strip()
        if not part_no:
            continue

        mrp_val = float(str(row.get("MRP", 0)).replace(",", "") or 0)

        db.execute(text("""
            INSERT INTO order_items (
                part_no, description, mrp, hsn
            ) VALUES (
                :part_no, :description, :mrp, :hsn
            )
        """), {
            "part_no": part_no,
            "description": row.get("PART DESC", ""),
            "mrp": mrp_val,
            "hsn": row.get("HSN", "")
        })

    db.commit()
    db.close()

    return {"message": "Full order data saved ✅"}
@app.get("/get_order_items")
def get_order_items(request: Request):
    active_store = get_active_store(request)
    db = SessionLocal()

    rows = db.execute(text("SELECT * FROM order_items WHERE store = :store"), {"store": active_store}).fetchall()
    result = []

    for r in rows:

        stock = db.query(Product).filter(
            Product.part_no == r.part_no,
            Product.store == active_store
        ).first()

        result.append({
            "part_no": r.part_no,
            "description": r.description,
            "rate": r.mrp,
            "stock": stock.quantity if stock else 0,
            "qty": stock.quantity if stock else 0,   # ✅ real qty
            "discount": 0,
            "status": "OUT OF STOCK" if (not stock or stock.quantity == 0) else "AVAILABLE"
        })

    # ✅ SORT AFTER LOOP
    result = sorted(result, key=lambda x: x["stock"] != 0)

    db.close()
    return result

from pydantic import BaseModel
from typing import List

class ChatMessage(BaseModel):
    role: str
    text: str

class ChatbotRequest(BaseModel):
    message: str
    history: List[ChatMessage] = []

@app.post("/chatbot/query")
async def chatbot_query(request: Request, data: ChatbotRequest):
    if not request.session.get("user"):
        return {"response": "⚠️ Unauthorized: Please log in to access the chatbot."}
    
    active_store = get_active_store(request)
    response_text = query_chatbot(data.message, [h.dict() for h in data.history], store_name=active_store)
    return {"response": response_text}

import os

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=int(os.environ.get("PORT", 8000))
    )


   
